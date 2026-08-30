#!/usr/bin/env python
"""
a500_backtest.py — A500 选股轻量回测

从已生成的 XLSX 报告中读取 Top 10，计算持有 N 日收益率。
不重跑选股流程，仅做前向收益统计。

用法:
    python a500_backtest.py                          # 回测最新报告
    python a500_backtest.py --date 2026-04-30        # 指定日期
    python a500_backtest.py --cron                   # 定时任务模式（回测上月所有报告）
    python a500_backtest.py --top 20 --periods 5,10,21  # 自定义参数

输出:
    a500_backtest_results.csv    — 逐笔收益明细
    a500_backtest_summary.csv    — 汇总统计
"""

import sys, os, json, time
from date_utils import date_to_str, parse_date_to_datetime
from datetime import datetime, timedelta
from collections import defaultdict
import pandas as pd
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data_manager import DataManager

# ============================================================
# 配置
# ============================================================
REPORT_DIR = "D:/常用文件/股票池推荐股"
OUTPUT_DIR = "D:/常用文件/股票池推荐股/股票池每月回测统计"
DEFAULT_TOP_N = 10
DEFAULT_PERIODS = [5, 10, 21]  # 5交易日≈1周, 10交易日≈2周, 21交易日≈1月

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ============================================================
# 1. 读取报告
# ============================================================

def find_reports(date_str: str = None) -> list:
    """找到 XLSX 报告文件列表

    Args:
        date_str: 指定日期 YYYY-MM-DD，None 则找最新的

    Returns:
        [(filepath, report_date_str), ...]
    """
    files = sorted(
        [f for f in os.listdir(REPORT_DIR) if f.startswith('扫描汇总_') and f.endswith('.xlsx')],
        reverse=True
    )

    if not files:
        print(f"未找到报告文件 (目录: {REPORT_DIR})")
        return []

    if date_str:
        matched = [f for f in files if date_str in f]
        if matched:
            files = matched
        else:
            print(f"未找到日期 {date_str} 的报告，使用最新报告")

    reports = []
    for f in files:
        # 提取日期: 扫描汇总_2026-04-30_1226.xlsx → 2026-04-30
        base = f.replace('扫描汇总_', '').replace('.xlsx', '')
        parts = base.split('_')
        if len(parts) >= 1:
            report_date = parts[0]  # YYYY-MM-DD
            reports.append((os.path.join(REPORT_DIR, f), report_date))

    return reports


def load_top_stocks(report_path: str, top_n: int = DEFAULT_TOP_N) -> list:
    """从 XLSX 中读取 Top N 股票

    Returns:
        [{'code': str, 'name': str, 'entry_price': float, 'grade': str, ...}, ...]
    """
    import openpyxl
    wb = openpyxl.load_workbook(report_path, read_only=True)
    ws = wb['Top10'] if 'Top10' in wb.sheetnames else wb.worksheets[0]

    stocks = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳过表头
        if i > top_n:
            break

        # 列: 排名(0), 代码(1), 名称(2), 综合分(3), 等级(4), 技术分(5),
        #     基本面分(6), 消息分(7), 仓位(8), 现价(9), 模式(10), 买点类型(11)
        try:
            stocks.append({
                'rank': int(row[0]),
                'code': str(row[1]).strip(),
                'name': str(row[2]).strip(),
                'entry_price': float(row[9]),
                'composite': float(row[3]),
                'grade': str(row[4]).strip() if row[4] else '',
                'tech_score': float(row[5]),
                'fund_score': float(row[6]),
                'news_score': float(row[7]),
                'pattern': str(row[10])[:40] if row[10] else '',
                'buy_type': str(row[11]).strip() if row[11] else '',
            })
        except (IndexError, ValueError, TypeError) as e:
            continue

    wb.close()
    return stocks


# ============================================================
# 2. 收益计算
# ============================================================

def calc_forward_returns(dm: DataManager, code: str, entry_date: str,
                         entry_price: float, periods: list) -> dict:
    """计算买入后的前向收益率

    优先用本地 K 线缓存，缓存过期时用 investoday MCP 获取实时行情。
    """
    result = {f'fwd_{p}d': None for p in periods}

    # 判断是否有前向数据：如果报告日期距今已超过最长周期，才算前向收益
    try:
        entry_dt = datetime.strptime(entry_date, "%Y-%m-%d")
    except:
        return result

    days_since = (datetime.now() - entry_dt).days
    if days_since < max(periods) * 1.4:  # 交易日≈自然日*1.4
        # 数据不足，标记为 pending
        return {f'fwd_{p}d': None for p in periods}

    # 取报告日后数据
    end_date = (entry_dt + timedelta(days=max(periods) * 2)).strftime("%Y-%m-%d")
    try:
        df = dm.get_klines(code, 'daily', entry_date, end_date)
    except Exception:
        return result

    if df.empty or len(df) < max(periods):
        return result

    # 找到 entry_date 位置
    df = df.reset_index(drop=True)
    entry_idx = None
    for i, row in df.iterrows():
        if str(row['date'])[:10] >= entry_date:
            entry_idx = i
            break

    if entry_idx is None:
        return result

    max_idx = len(df) - 1
    for p in periods:
        target_idx = entry_idx + p
        if target_idx <= max_idx:
            exit_price = float(df.iloc[target_idx]['close'])
            ret = (exit_price - entry_price) / entry_price
            result[f'fwd_{p}d'] = round(ret, 4)

    return result


# ============================================================
# 3. 主流程
# ============================================================

def run_backtest(report_path: str, report_date: str, dm: DataManager,
                 top_n: int = DEFAULT_TOP_N, periods: list = DEFAULT_PERIODS):
    """对单份报告进行回测"""
    stocks = load_top_stocks(report_path, top_n)
    if not stocks:
        print(f"  {report_date}: 无股票数据")
        return [], None

    print(f"  {report_date}: {len(stocks)} 只股票, 现价基准日={report_date}")

    records = []
    for s in stocks:
        fwd = calc_forward_returns(dm, s['code'], report_date, s['entry_price'], periods)
        record = {
            'report_date': report_date,
            'rank': s['rank'],
            'code': s['code'],
            'name': s['name'],
            'entry_price': s['entry_price'],
            'composite': s['composite'],
            'grade': s['grade'],
            'tech_score': s['tech_score'],
            'fund_score': s['fund_score'],
            'news_score': s['news_score'],
            'pattern': s['pattern'],
            'buy_type': s['buy_type'],
            **fwd,
        }
        records.append(record)

    # 汇总
    df = pd.DataFrame(records)
    summary = {'report_date': report_date, 'n_stocks': len(records)}
    for p in periods:
        col = f'fwd_{p}d'
        valid = df[col].dropna()
        if len(valid) > 0:
            summary[f'avg_{p}d'] = valid.mean()
            summary[f'win_{p}d'] = (valid > 0).sum() / len(valid)
            summary[f'median_{p}d'] = valid.median()
            summary[f'max_{p}d'] = valid.max()
            summary[f'min_{p}d'] = valid.min()
        else:
            summary[f'avg_{p}d'] = None
            summary[f'win_{p}d'] = None

    return records, summary


def main():
    args = sys.argv[1:]
    cron_mode = '--cron' in args
    date_str = None
    top_n = DEFAULT_TOP_N
    periods = DEFAULT_PERIODS

    for i, a in enumerate(args):
        if a == '--date' and i + 1 < len(args):
            date_str = args[i + 1]
        elif a == '--top' and i + 1 < len(args):
            top_n = int(args[i + 1])
        elif a == '--periods' and i + 1 < len(args):
            periods = [int(x) for x in args[i + 1].split(',')]

    # === cron 模式：检查上游依赖 ===
    if cron_mode:
        from cron_utils import FlagSignals, CronLogger
        logger = CronLogger("a500_backtest")
        logger.info("cron 模式: 检查上游依赖...")

        # 检查最近一次 a500_scan_done 标记
        latest = FlagSignals.get_latest("a500_scan_done")
        if not latest:
            logger.warn("未找到 a500_scan_done 标记，跳过回测")
            return
        flag_data = FlagSignals.read("a500_scan_done",
                                      os.path.basename(latest).replace('a500_scan_done_', '').replace('.flag', ''))
        logger.info(f"上游标记: {os.path.basename(latest)} "
                     f"(候选{flag_data.get('candidates', '?')}只, "
                     f"评分{flag_data.get('scored', '?')}只, "
                     f"耗时{flag_data.get('elapsed_s', '?')}s)")

    dm = DataManager()

    # 找报告
    if cron_mode:
        # 定时任务：回测上个月的所有报告
        now = datetime.now()
        last_month = now - timedelta(days=30)
        date_str = last_month.strftime("%Y-%m")
        print(f"[Cron] 回测 {date_str} 月份报告...")

    reports = find_reports(date_str)
    if not reports:
        print("无报告可回测")
        return

    print(f"找到 {len(reports)} 份报告, Top {top_n}, 周期={periods}\n")

    all_records = []
    all_summaries = []

    for report_path, report_date in reports:
        records, summary = run_backtest(report_path, report_date, dm, top_n, periods)
        if records:
            all_records.extend(records)
        if summary:
            all_summaries.append(summary)

    if not all_records:
        print("无回测数据")
        return

    # ── 保存 ──
    df_raw = pd.DataFrame(all_records)
    raw_path = os.path.join(OUTPUT_DIR, "a500_backtest_results.csv")
    df_raw.to_csv(raw_path, index=False, encoding='utf-8-sig')

    df_sum = pd.DataFrame(all_summaries)
    sum_path = os.path.join(OUTPUT_DIR, "a500_backtest_summary.csv")
    df_sum.to_csv(sum_path, index=False, encoding='utf-8-sig')

    # ── 报告 ──
    print(f"\n{'='*70}")
    print(f"A500 选股回测报告")
    print(f"{'='*70}")
    print(f"回测报告数: {len(reports)}")
    print(f"总记录: {len(all_records)} 条")
    print()

    # 按等级分组
    for p in periods:
        col = f'fwd_{p}d'
        print(f"── T+{p}交易日 (持有{p}天) ──")
        for grade in ['A', 'B', 'C', 'D']:
            gdf = df_raw[(df_raw['grade'].fillna('').str.startswith(grade)) & df_raw[col].notna()]
            if len(gdf) > 0:
                print(f"  {grade}级 (n={len(gdf):>2}): 均值={gdf[col].mean():>+7.2%}  "
                      f"胜率={ (gdf[col] > 0).sum() / len(gdf):.0%}  中位={gdf[col].median():>+7.2%}")
        print()

    # 每份报告的月度统计
    print("── 报告月度表现 ──")
    for s in all_summaries:
        parts = []
        for p in periods:
            avg = s.get(f'avg_{p}d')
            wr = s.get(f'win_{p}d')
            if avg is not None:
                parts.append(f"T+{p}d: {avg:>+6.1%} (wr={wr:.0%})")
        print(f"  {s['report_date']}: {' | '.join(parts)}")

    print(f"\n原始数据: {raw_path}")
    print(f"汇总统计: {sum_path}")


if __name__ == "__main__":
    main()
