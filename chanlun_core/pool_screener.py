#!/usr/bin/env python
"""pool_screener.py — A500 股票池智能筛选主控

Phase 1: 轻量技术面扫描 → 筛选有买点/潜在买点的股票
Phase 2: 三维深度评估（技术+基本面+消息面）
Phase 3: 生成报告（HTML技术分析 + MD评分报告 + Excel/MD总表）

用法:
    python pool_screener.py                          # 全流程
    python pool_screener.py --phase1-only            # 仅扫描
    python pool_screener.py --from-cache             # 从缓存加载Phase1结果继续
    python pool_screener.py --test N                 # 仅跑前N只（测试用）
"""

import sys, os, json, time, shutil, threading
import socket
socket.setdefaulttimeout(60)  # 全局 socket 超时 60s，防止 AKShare 请求挂死
from date_utils import date_to_str, parse_date_to_datetime
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import numpy as np
from dotenv import load_dotenv

# ── 加载 ~/.hermes/.env 环境变量 ──
# 背景进程的 $HOME 可能被 profile 覆盖，需从真实 home 路径加载 .env 文件
_real_home = os.path.expanduser("~")
_hermes_home = os.environ.get("HERMES_HOME", "")
if _hermes_home and os.path.isdir(_hermes_home):
    # HERMES_HOME 指向 profile 目录（如 .../profiles/commander），父目录是真实 HERMES_HOME
    _parent = os.path.dirname(os.path.dirname(_hermes_home.rstrip("/")))
    _dotenv = os.path.join(_parent, ".env")
elif os.path.exists(os.path.join(_real_home, ".hermes", ".env")):
    _dotenv = os.path.join(_real_home, ".hermes", ".env")
else:
    _dotenv = ""
if _dotenv and os.path.exists(_dotenv):
    load_dotenv(_dotenv)
    _tavily = os.environ.get("TAVILY_API_KEY", "")
    _metaso = os.environ.get("METASO_API_KEY", "")
    print(f"[Env] 已加载 {_dotenv} (TAVILY={'有' if _tavily else '无'}, METASO={'有' if _metaso else '无'})")
else:
    print("[Env] 未找到 .env 文件，API 密钥仅从系统环境变量读取")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import baostock_utils
from file_utils import safe_read_json, safe_write_json
from config_loader import (
    W_TECH, W_FUND, W_ALPHA, W_NEWS,
    A500_SCORE_THRESHOLD, A500_COMPOSITE_THRESHOLD,
    A500_TOP_N_REPORT, A500_BATCH_COUNT, A500_BATCH_PAUSE, A500_NEWS_TOP_N,
    MANUAL_BLACKLIST,  # v5.3.1(F2): veto 检查需人工黑名单
    TECH_BUY_THRESHOLD,  # v5.3.2(D-5/X1): 报告脚注引用建仓线
    SELL_SIGNAL_SUPPRESS_DAYS,  # v5.3.3(E-1): 买卖冲突仲裁窗口
)
from data_manager import DataManager

# 线程锁：Baostock 连接非线程安全
# v5.3.3(F-2): 统一使用 baostock_utils.BS_SESSION_LOCK——原独立 _BS_LOCK
# 只护住本文件的 fallback, akshare_fundamental 内部 Baostock 调用未受保护,
# 4线程并发 session 串包(川投 PE=0 事故)。别名保留兼容既有引用。
import baostock_utils as _bsu
_BS_LOCK = _bsu.BS_SESSION_LOCK

from generate_analysis import ChanLunAnalyzer, RecursiveTimingSystem, HTMLVisualizer
from composite_scorer import compute_3d_score, position_reason, Score3D
from quick_fundamental import (
    classify_by_industry, calculate_fundamental_score,
    classify_stock_type, get_fundamentals,
)
from akshare_fundamental import get_fundamentals_akshare

# ============================================================
# 配置
# ============================================================

OUTPUT_BASE = "D:/常用文件/股票池推荐股"
CACHE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".scanner_cache.json")
SCORE_THRESHOLD = A500_SCORE_THRESHOLD          # Phase 1 最低分（≥3 标准筛选，约 20-40 只进 Phase2）
COMPOSITE_THRESHOLD = A500_COMPOSITE_THRESHOLD     # 综合分最低门槛（低于此分不推荐）
TOP_N_REPORT = A500_TOP_N_REPORT            # 报告中最多展示前N只

BATCH_COUNT = A500_BATCH_COUNT             # Phase 2 分几批执行（避免 Baostock 限流）
BATCH_PAUSE = A500_BATCH_PAUSE            # 批次间暂停秒数
NEWS_TOP_N = A500_NEWS_TOP_N             # 消息面仅扫描技术+基本面综合 Top N（省 Tavily 调用）

# 三维权重（与 composite_scorer 一致，从 config_loader 继承）
# W_TECH, W_FUND, W_NEWS 已在文件顶部从 config_loader 导入

# ============================================================
# 市场环境仓位上限（从 market_regime.py 输出读取）
# ============================================================

def get_regime_position_cap() -> float:
    """读取 market_regime.py 输出的最新仓位上限，默认 1.0（不限制）"""
    import csv
    regimes_csv = os.path.join(os.path.dirname(os.path.abspath(__file__)), "regimes.csv")
    if not os.path.exists(regimes_csv):
        return 1.0
    
    try:
        with open(regimes_csv, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        if rows:
            latest = rows[-1]
            cap = float(latest.get('position_cap', 1.0))
            regime = latest.get('regime', 'unknown')
            print(f"[大盘仓位上限] {regime} → 个股最高 {cap*100:.0f}%")
            return cap
    except Exception as e:
        print(f"[大盘仓位上限] 读取失败: {e}，默认不限制")
    
    return 1.0

# 全局缓存，只读一次
_REGIME_POSITION_CAP = None

def _get_cached_cap() -> float:
    global _REGIME_POSITION_CAP
    if _REGIME_POSITION_CAP is None:
        _REGIME_POSITION_CAP = get_regime_position_cap()
    return _REGIME_POSITION_CAP

# ============================================================
# Phase 1: 技术面扫描（调用 pool_scanner）
# ============================================================

def run_phase1(test_n: int = 0):
    """运行 Phase 1 扫描，返回 candidates 列表"""
    import pool_scanner

    # 临时限制股票池（测试用）
    _orig_load_pool = pool_scanner.load_a500_pool
    _orig_threshold = pool_scanner.SCORE_THRESHOLD
    _orig_cache_path = pool_scanner.CACHE_PATH
    if test_n > 0:
        pool_scanner.load_a500_pool = lambda: _orig_load_pool()[:test_n]
        pool_scanner.SCORE_THRESHOLD = 0  # 测试时不设阈值
        # v4.2 修复：--test N 写独立测试缓存，避免污染生产缓存
        # 原实现 test 模式也写 CACHE_PATH → 后续 --from-cache 基于残缺数据
        pool_scanner.CACHE_PATH = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f".scanner_cache_test{test_n}.json"
        )

    try:
        result = pool_scanner.main()
    finally:
        # 恢复原值（防止全局污染）
        pool_scanner.load_a500_pool = _orig_load_pool
        pool_scanner.SCORE_THRESHOLD = _orig_threshold
        pool_scanner.CACHE_PATH = _orig_cache_path

    candidates = result['candidates']

    print(f"\n[Phase 1] 候选股: {len(candidates)} 只")
    return candidates, result


def load_phase1_cache():
    """从缓存加载 Phase 1 结果"""
    if not os.path.exists(CACHE_PATH):
        print("[Phase 1] 缓存不存在，运行全量扫描...")
        return run_phase1()

    data = safe_read_json(CACHE_PATH)

    candidates = [r for r in data.get('candidates', []) if r['score'] >= SCORE_THRESHOLD]
    print(f"[Phase 1] 从缓存加载: {len(candidates)} 只候选 (扫描时间: {data.get('scan_time')})")
    return candidates, data


# ============================================================
# Phase 2: 三维深度评估
# ============================================================

def _buy_level_from_type(buy_type: str) -> int:
    """从买点类型字符串提取缠论级别（1=一买 2=二买 3=三买 0=其他）。
    用于综合评分的仓位调整（一买升档/三买降档）。
    v5.3.1(F1): 委托 composite_scorer 公共实现（供 resocre/ff_rescore 复用）。"""
    from composite_scorer import buy_level_from_type
    return buy_level_from_type(buy_type)


def scan_news(code: str, name: str, provider: str = "tavily") -> tuple:
    """多源消息面扫描 — 委托给 news_scanner.scan_news()"""
    from news_scanner import scan_news as _ns_scan_news
    return _ns_scan_news(code, name)


def _write_news_fallback(code: str, name: str, error_detail: str):
    """委托给 news_scanner._write_news_fallback()"""
    from news_scanner import _write_news_fallback as _ns_write_fallback
    return _ns_write_fallback(code, name, error_detail)


def list_news_fallbacks() -> list:
    """委托给 news_scanner.list_news_fallbacks()"""
    from news_scanner import list_news_fallbacks as _ns_list
    return _ns_list()


def run_phase2(candidates: list, batch_count: int = 5, batch_pause: int = 20) -> list:
    """对每只候选股运行三维评估（分批执行避免 Baostock 限流）

    Args:
        candidates: Phase 1 筛选的候选股列表
        batch_count: 分几批 (默认5批)
        batch_pause: 批次间暂停秒数 (默认20秒)
    """
    import math
    batch_size = max(1, math.ceil(len(candidates) / batch_count))
    scored = []

    # Phase 2 开始时重置 Baostock session（后续复用，不再每只重连）
    baostock_utils.logout()
    baostock_utils.ensure_login()

    for bi in range(batch_count):
        start = bi * batch_size
        end = min(start + batch_size, len(candidates))
        if start >= len(candidates):
            break
        batch = candidates[start:end]

        print(f"\n--- 批次 {bi+1}/{batch_count} ({len(batch)} 只, {start+1}-{end}/{len(candidates)}) ---")

        batch_results = _process_batch(batch, global_offset=start, skip_news=True)
        scored.extend(batch_results)

        # 每批完成后增量保存（防崩溃丢数据）
        _save_phase2_results(scored, suffix=f"_batch{bi+1}")

        # 批次间暂停（最后一批不用）
        remaining = len(candidates) - end
        if remaining > 0:
            print(f"  批次 {bi+1} 完成，{remaining} 只待处理，暂停 {batch_pause}s...")
            time.sleep(batch_pause)

    # 按技术+基本面综合分降序（消息面暂时=50中性）
    scored.sort(key=lambda s: -s['composite'])
    print(f"\n[Phase 2] 技术+基本面评分完成: {len(scored)} 只有效评分")

    # ── 消息面补扫：仅对 Top N（并行版）──
    top_n = min(NEWS_TOP_N, len(scored))
    print(f"\n[Phase 2] 消息面补扫 Top {top_n}（8线程并行）...")
    news_updated = [0]  # 用 list 模拟 nonlocal

    def _update_news(i):
        """更新单只股票的消息面评分"""
        s = scored[i]
        code, name = s['code'], s['name']
        try:
            news_score, news_detail = scan_news(code, name)
            s['news_score'] = round(news_score, 1)
            s['news_detail'] = news_detail
            result_3d = compute_3d_score(
                tech_score=s['tech_score'],
                fund_score=s['fund_score'],
                alpha_score=s.get('alpha_score', 50.0),
                news_score=news_score,
                w_tech=W_TECH, w_fund=W_FUND, w_alpha=W_ALPHA, w_news=W_NEWS,
                code=code, name=name,
                news_detail=news_detail,
                resonance_penalty=True,
                buy_level=_buy_level_from_type(s.get('buy_type', '')),
                # v5.3.1(F2): 消息面重算点同样接通 severe 链
                risk_reasons=(s.get("risk_reasons") or []) + (s.get("severe_reasons") or []),
                manual_blacklist=MANUAL_BLACKLIST,
                # v5.3.3(E-1/E-2): 从 phase2 标志透传
                recent_top_sell=bool(s.get('sell_conflict') or s.get('suppressed_by_sell')),
                observational=bool(s.get('observational')),
            )
            s['composite'] = result_3d.composite
            s['grade'] = result_3d.grade
            # v5.3.1(F1/F8): position 与 position_pct 必须同源——cap 后统一生成,
            # 否则报告(读 pct)与存储(float)口径分裂(A级印15%事故)
            _pos_capped = min(result_3d.position, _get_cached_cap())
            s['position'] = _pos_capped
            s['position_pct'] = f"{_pos_capped*100:.0f}%"
            s['can_buy'] = result_3d.can_buy
            s['reason'] = position_reason(result_3d)
            s['resonance'] = result_3d.components.get('resonance_penalty_applied', False)
            s['sell_conflict'] = result_3d.components.get('sell_conflict', False)
            s['observational'] = result_3d.components.get('observational', False)
            news_updated[0] += 1
            return f"{code} {name} news={news_score:.0f} → {result_3d.composite:.0f}({result_3d.grade})"
        except Exception as e:
            return f"{code} {name} 消息面失败: {e}"

    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=4) as pool:
        nf = {pool.submit(_update_news, i): i for i in range(top_n)}
        for f in as_completed(nf):
            i = nf[f]
            try:
                msg = f.result(timeout=120)
                print(f"  [{i+1}/{top_n}] {msg}")
            except Exception as e:
                print(f"  [{i+1}/{top_n}] 线程异常: {e}")

    print(f"  消息面补扫完成: {news_updated[0]}/{top_n}")

    # 按最终综合分重排
    scored.sort(key=lambda s: -s['composite'])
    print(f"\n[Phase 2] 最终完成: {len(scored)} 只有效评分")

    # 保存真实评分到 JSON（供 full_rescore.py 读取）
    _save_phase2_results(scored)
    return scored


def _save_phase2_results(scored: list, suffix: str = ""):
    """保存 Phase 2 真实技术/基本面评分到 JSON"""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f".phase2_results{suffix}.json")
    # v5.3.3(G-3): fund_data 不再整体剥离——保留报告层所需的可序列化子集。
    # 历史: 整体剥离导致 ff_rescore→generate_reports 链的季报点评表退化为
    # 仅 ROE 一行(其余指标全依赖 fund_data)。analyzer 仍剥离(不可序列化)。
    _FD_KEYS = ('profitability', 'growth', 'health', 'valuation',
                'quarterly_profits', 'multi_year_data', 'data_date', 'industry')
    clean = []
    for s in scored:
        d = {k: v for k, v in s.items() if k != 'analyzer'}
        fd = s.get('fund_data')
        if isinstance(fd, dict):
            d['fund_data'] = {k: fd.get(k) for k in _FD_KEYS if k in fd}
        clean.append(d)
    safe_write_json(path, clean)
    if suffix:
        print(f"  增量保存: {path} ({len(clean)} 只)")


def _process_batch(batch: list, global_offset: int = 0, skip_news: bool = False) -> list:
    """并行处理一批候选股（线程池，Baostock 加锁串行，HTTP API 并行）"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    scored = []
    _lock = threading.Lock()  # 保护 scored.append 的顺序（可选）
    
    def _work(i, c):
        """线程内执行单只股票分析"""
        local = []
        idx = global_offset + i
        print(f"  [{idx}] {c['code']} {c['name']}...", end=" ", flush=True)
        _process_one_stock(i, c, local, skip_news=skip_news)
        return local
    
    n_workers = min(4, len(batch)) if len(batch) >= 4 else len(batch)
    with ThreadPoolExecutor(max_workers=n_workers) as pool:
        futures = {pool.submit(_work, i, c): i for i, c in enumerate(batch)}
        for f in as_completed(futures):
            try:
                result = f.result(timeout=300)
                scored.extend(result)
            except Exception as e:
                print(f"  线程异常: {e}")
    
    return scored


def _process_one_stock(i: int, c: dict, batch_output: list, skip_news: bool = False):
    """处理单只股票的三维评分，结果追加到 batch_output"""
    code, name = c['code'], c['name']

    # ── 8项风控过滤 ──
    from risk_filter import check_risk
    from config_loader import MANUAL_BLACKLIST
    blocked, risk_reasons = check_risk(code, name, MANUAL_BLACKLIST)
    if blocked:
        print(f"⛔ 风控排除: {' | '.join(risk_reasons)}")
        return
    # v5.3.1(F2): 风控理由随候选传递, 供所有 compute_3d_score 调用点的
    # apply_veto 消费(severe -15分+限轻仓)。此前从未传入→severe 链断裂,
    # 解禁预警/监管处罚只躺在字段里无人扣分。
    c["risk_reasons"] = list(risk_reasons or [])

    # ── 风控增强（P1-2）：重点监控池（缓存快查）→ 否决；解禁预警 → severe/veto ──
    try:
        from risk_enhance import check_regulatory_risks
        reg = check_regulatory_risks(code)
        if reg["level"] == "veto":
            print(f"⛔ 风控增强否决: {' | '.join(reg['reasons'])}")
            return
        if reg["level"] == "severe":
            c.setdefault("severe_reasons", [])
            c["severe_reasons"].extend(reg["reasons"])
    except Exception as _re_e:
        print(f"  [风控增强] 检查失败({str(_re_e)[:40]})，跳过")

    try:
        # --- 技术面评分（Baostock 部分加锁） ---
        m30_analyzer = None
        with _BS_LOCK:
            dm = DataManager()
            df = dm.get_klines(code, 'daily', '2024-01-01', datetime.now().strftime("%Y-%m-%d"))
            if df.empty or len(df) < 60:
                print("K线不足")
                return

            klines = dm.to_json_list(df)
            analyzer = ChanLunAnalyzer('daily', min_bi_klines=5)
            analyzer.analyze(klines)

            # ── v5.3.1(F5): 接入真实 30min 共振——用户拍板恢复名义五维评分。
            # 共振块(±5天买点+8/类型一致+5/时间同步+3/背驰+2/中枢同向+2)此前
            # 因生产恒传 m30_analyzer=None 而成死代码, 全员白拿基础5分。
            # 冷缓存时 30min 增量链路已由 P0-1 修复保证安全; 失败降级为
            # m30_analyzer=None(仅日线, 共振=基础5分), 不阻塞主流程。
            try:
                _m30_df = dm.get_klines(code, '30min', '2024-06-01')
                if not _m30_df.empty and len(_m30_df) >= 200:
                    m30_analyzer = ChanLunAnalyzer('30min', min_bi_klines=5)
                    m30_analyzer.analyze(dm.to_json_list(_m30_df))
            except Exception as _m30_e:
                print(f"[30min] {code} 共振数据不可用({str(_m30_e)[:50]}), 仅日线评分")

        # 找到最近买点（klines[-1] 是 dict，用 ['date'] 访问）
        buy_points = [p for p in analyzer.buy_sell_points if p.type == 'buy']
        recent_buy = None
        last_kline = klines[-1] if klines else {}
        if buy_points and last_kline:
            last_dt = datetime.strptime(str(last_kline.get('date', ''))[:10], "%Y-%m-%d")
            for bp in sorted(buy_points, key=lambda p: str(p.date), reverse=True):
                bp_dt = datetime.strptime(date_to_str(bp.date), "%Y-%m-%d")
                if (last_dt - bp_dt).days <= 120:
                    recent_buy = bp
                    break

        # ── fallback: scanner 缓存买点代理（修复反转后三买等无标准买点场景）──
        # 天山铝业等买点：scanner 检测到的三买在重建分析器后 >120 天被筛掉，
        # 但 scanner 缓存中有 buy_type/buy_price/buy_date 可供参考。
        if recent_buy is None and c.get('score', 0) >= 3 and c.get('buy_type', ''):
            from types import SimpleNamespace
            _bp = SimpleNamespace()
            _bp.price = float(c.get('buy_price', 0) or last_kline.get('close', 0))
            _bp.level = {'一买': 1, '二买': 2, '三买': 3,
                         '反转后三买': 3, '反转后类二买': 2}.get(c.get('buy_type', ''), 0)
            _bp.confirmed = True
            try:
                _bp.date = datetime.strptime(str(c.get('buy_date', ''))[:10], "%Y-%m-%d")
            except Exception:
                _bp.date = None
            _bp.reason = ''
            # v5.3.1(M3): 代理买点透传 scanner 缓存的中枢边界作为参照——
            # Phase1 时点的中枢即三买发生时的突破基准(Phase2 重算时点的
            # zhongshus[-1] 可能已错位)。反转路径缓存 zg/zd 即 display 基准。
            _bp.ref_zg = c.get('zg')
            _bp.ref_zd = c.get('zd')
            _bp.multilevel_confirmation = {'confidence_score': 0, 'm30_confirmation': False}
            recent_buy = _bp
            print(f"(用扫描器买点代理:{c.get('buy_type','')}@{c.get('buy_date','')})", end=" ")

        # ── v5.3.3(E-1): 二次防御——Phase2 数据比扫描更新, 一卖/二卖可能新出现 ──
        from config_loader import SELL_SIGNAL_SUPPRESS_DAYS
        _recent_top_sell = False
        try:
            _last_dt = datetime.strptime(str(klines[-1].get('date', ''))[:10], "%Y-%m-%d")
            for _sp in analyzer.buy_sell_points:
                if getattr(_sp, 'type', '') == 'sell' and getattr(_sp, 'level', 0) in (1, 2):
                    try:
                        _sd = datetime.strptime(date_to_str(_sp.date), "%Y-%m-%d")
                    except Exception:
                        continue
                    if 0 <= (_last_dt - _sd).days <= SELL_SIGNAL_SUPPRESS_DAYS:
                        _recent_top_sell = True
                        print(f"[E-1] {code} 近{SELL_SIGNAL_SUPPRESS_DAYS}日内有"
                              f"{'一卖' if getattr(_sp,'level',0)==1 else '二卖'}"
                              f"@{date_to_str(_sp.date)}, 买入压制", end=" ")
                        break
        except Exception:
            pass

        # 使用 validate_tech_score 做精细评分（带 fallback）
        from validate_tech_score import compute_technical_score
        tech_degraded = False
        try:
            # v5.3.1(F5): 传入真实 30min 分析器, 共振20分从死代码变为实评
            tech_result = compute_technical_score(analyzer, m30_analyzer, recent_buy)
            tech_score = tech_result.get('tech_score', c['score'] * 20)
        except Exception as te:
            # 评分失败 → 用扫描评分估算
            tech_score = c['score'] * 20
            tech_degraded = True
            print(f"(技术分降级)", end=" ")

        # --- 基本面评分（AKShare 优先 → Baostock fallback）---
        fund_data = None
        fund_source = ""
        fund_degraded = False
        fund_degraded_reason = ""

        # 1. AKShare (数据更丰富，无 session 冲突)
        try:
            fund_data = get_fundamentals_akshare(code)
            if fund_data.get("error") or fund_data.get("profitability", {}).get("roeAvg") is None:
                fund_degraded = True
                fund_degraded_reason = f"AKShare降级:{str(fund_data.get('error',''))[:30]}"
                print(f"({fund_degraded_reason})", end=" ")
                fund_data = None
            else:
                fund_source = "akshare"
        except Exception as fe:
            fund_degraded = True
            fund_degraded_reason = f"AKShare异常:{str(fe)[:30]}"
            print(f"({fund_degraded_reason})", end=" ")
            fund_data = None

        # 2. Baostock fallback（v4.2 加锁：Baostock 是连接级会话，
        # 原实现4线程并发调用 get_fundamentals 会跨线程串包——A股票的
        # 财务数据可能被 B股票覆盖。与技术面扫描共用 _BS_LOCK）
        if fund_data is None:
            try:
                with _BS_LOCK:
                    fund_data = get_fundamentals(code)
                if fund_data.get("error"):
                    fund_data = None
                else:
                    fund_source = "baostock"
            except Exception as fe:
                fund_degraded = True
                fund_degraded_reason += f" | Baostock异常:{str(fe)[:30]}"
                print(f"(Baostock基本面异常:{str(fe)[:30]})", end=" ")
                fund_data = None

        if fund_data is None:
            fund_score = 50  # 数据缺失默认中性
            roe, np_margin, industry, pe, pb = 0, 0, "", 0, 0
            # v4.2 修复：双基本面源失败时初始化 fund_score_obj 空对象，
            # 否则 L481-485 引用未定义变量 → NameError → 整只股票被外层
            # except 吞掉静默消失（设计的中性50降级失效，股票从结果消失）
            fund_score_obj = {}
            fund_degraded = True
            if not fund_degraded_reason:
                fund_degraded_reason = "双源均失败"
            print(f"(基本面全部降级)", end=" ")
        else:
            fund_score_obj = calculate_fundamental_score(fund_data, multi_year_data=fund_data.get('multi_year_data'))
            fund_score = fund_score_obj.get('total_score', 50)
            roe = fund_data.get('profitability', {}).get('roeAvg', 0) or 0
            np_margin = fund_data.get('profitability', {}).get('npMargin', 0) or 0
            industry = fund_data.get('industry', '')
            # v5.3.3(F-2): PE/PB 缺失保持 None(报告显示—), 不再落成 0 假值
            pe = fund_data.get('valuation', {}).get('peTTM')
            pb = fund_data.get('valuation', {}).get('pbMRQ')
            if fund_source == "akshare":
                print(f"[AKShare:{fund_score}]", end=" ")

        # --- 研报评级微调 (-2 ~ +5分) ---
        if fund_data is not None:
            try:
                from akshare_fundamental import scan_research_reports
                rr_delta, rr_detail = scan_research_reports(code)
                if rr_delta != 0:
                    fund_score = max(15, min(100, fund_score + rr_delta))
                    print(f" {rr_detail}", end="")
            except Exception:
                pass  # 研报扫描失败不影响主流程

        # --- 消息面评分 ---
        if skip_news:
            news_score = 50.0
            news_detail = "跳过(非Top30)"
        else:
            news_score, news_detail = scan_news(code, name)

        # --- 综合评分（四维 + Veto） ---
        alpha_score = c.get("alpha_score", 50.0)
        result_3d = compute_3d_score(
            tech_score=tech_score,
            fund_score=fund_score,
            alpha_score=alpha_score,
            news_score=news_score,
            w_tech=W_TECH, w_fund=W_FUND, w_alpha=W_ALPHA, w_news=W_NEWS,
            code=code, name=name,
            news_detail=news_detail,
            resonance_penalty=True,
            buy_level=_buy_level_from_type(c.get('buy_type', '')),
            # v5.3.1(F2): 接通 severe 链——risk_filter 理由 + 风控增强解禁预警
            risk_reasons=(c.get("risk_reasons") or []) + (c.get("severe_reasons") or []),
            manual_blacklist=MANUAL_BLACKLIST,
            # v5.3.3(E-1): 近期一卖/二卖压制（scanner标记或本机复查命中）
            recent_top_sell=_recent_top_sell or bool(c.get('suppressed_by_sell')),
            # v5.3.3(E-2): 观察型几何信号——入池但仓位封顶轻仓
            observational=bool(c.get('observational')),
        )

        batch_output.append({
            **c,
            "tech_score": round(tech_score, 1),
            "fund_score": round(fund_score, 1),
            "news_score": round(news_score, 1),
            "composite": result_3d.composite,
            "grade": result_3d.grade,
            # v5.3.1(F1/F8): cap 后同源生成 position/position_pct
            "position": min(result_3d.position, _get_cached_cap()),
            "position_pct": f"{min(result_3d.position, _get_cached_cap())*100:.0f}%",
            "can_buy": result_3d.can_buy,
            "reason": position_reason(result_3d),
            "resonance": result_3d.components.get('resonance_penalty_applied', False),
            # v5.3.3(E-1/E-2): 标志透传到 phase2 json, 供 resocre/ff_rescore 链传递
            "sell_conflict": result_3d.components.get('sell_conflict', False),
            "observational": result_3d.components.get('observational', False),
            "roe": round(roe, 2),
            "np_margin": round(np_margin, 2),
            # v5.3.3(F-2): None 保留为 None(报告层显示"—"), 不再 round 崩溃/落0
            "pe": round(pe, 2) if pe is not None else None,
            "pb": round(pb, 2) if pb is not None else None,
            "industry": industry,
            "news_detail": news_detail,
            "analyzer": analyzer,  # 保留分析器用于生成HTML
            "fund_data": fund_data,
            # v5.3.3(F-1/F-3): ROE口径与净利率失真标注, 报告层渲染
            "roe_basis": fund_score_obj.get('roe_basis', ''),
            "annual_roe": fund_score_obj.get('annual_roe'),
            "margin_note": fund_score_obj.get('margin_note', ''),
            "profitability_details": fund_score_obj.get('profitability_details', []),
            "valuation_degraded": bool(fund_data.get('valuation_degraded')) if isinstance(fund_data, dict) else False,
            "trend_correction": fund_score_obj.get('trend_correction', 0),
            "trend_correction_detail": fund_score_obj.get('trend_correction_detail', ''),
            "roe_std": fund_score_obj.get('roe_std'),
            "revenue_volatility": fund_score_obj.get('revenue_volatility'),
            "multi_year_data": fund_data.get('multi_year_data'),
            # v5.0.1 P2：数据降级标记（区分"差"和"没数据"）
            "data_degraded": {
                "tech": tech_degraded,
                "fund": fund_degraded,
                "fund_reason": fund_degraded_reason,
            },
        })
        print(f"综合{result_3d.composite:.0f}({result_3d.grade})")

    except Exception as e:
        print(f"失败: {e}")


# ============================================================
# Phase 3: 报告生成
# ============================================================

def generate_html_report(s: dict, output_dir: str):
    """生成缠论技术分析 HTML 报告（快速模式+潜在一买标签）"""
    code = s['code']
    name = s['name']
    dst = os.path.join(output_dir, f"{code}_chanlun.html")

    # 优先用 quick_html（含潜在一买标签）
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from quick_html import generate_html as qh_generate
        qh_result = qh_generate(code, name)
        if qh_result and not qh_result.get('error'):
            src = qh_result.get('html_path', '')
            if src and os.path.exists(src):
                import shutil
                shutil.copy2(src, dst)
                return dst
    except Exception:
        pass

    # fallback: 用内存 analyzer
    analyzer = s.get('analyzer')
    if not analyzer:
        return None
    try:
        visualizer = HTMLVisualizer(code, name, analyzer)
        visualizer.generate_html(dst)
        return dst
    except Exception as e:
        print(f"    HTML生成失败 {code}: {e}")
        return None


def _valuation_percentile(symbol, cur_pe=None, cur_pb=None):
    """v5.3.3(G-2): Baostock 7年日线 PE-TTM/PB 历史分位（stock-analysis 标准）。

    Returns: {'pe': {min,max,median,n,cur,pct}, 'pb': {...}} 或 None。
    持 BS_SESSION_LOCK(F-2 教训), 失败静默返 None——报告增强项不阻塞主流程。
    """
    try:
        import baostock as _bs
        from datetime import timedelta as _td
        end = datetime.now().strftime('%Y-%m-%d')
        start = (datetime.now() - _td(days=365 * 7)).strftime('%Y-%m-%d')
        with _bsu.BS_SESSION_LOCK:
            _, lg = _bsu.ensure_login()
            if lg is None or lg.error_code != "0":
                return None
            rs = _bs.query_history_k_data_plus(
                _bsu.to_bs_code(symbol), "date,peTTM,pbMRQ",
                start_date=start, end_date=end, frequency="d", adjustflag="3")
            dfv = rs.get_data()
        if dfv is None or dfv.empty or 'peTTM' not in getattr(dfv, 'columns', []):
            return None
        pes = pd.to_numeric(dfv['peTTM'], errors='coerce')
        pes = pes[(pes > 0) & (pes < 2000)].dropna()
        pbs = pd.to_numeric(dfv['pbMRQ'], errors='coerce')
        pbs = pbs[(pbs > 0) & (pbs < 300)].dropna()

        def _stats(series, cur):
            if series is None or len(series) == 0:
                return None
            st = {"min": round(float(series.min()), 1), "max": round(float(series.max()), 1),
                  "median": round(float(series.median()), 1), "n": int(len(series))}
            if cur is not None and cur > 0:
                c = float(cur)
                st["cur"] = round(c, 1)
                st["pct"] = round(float((series < c).mean()) * 100, 1)
            return st

        return {"pe": _stats(pes, cur_pe), "pb": _stats(pbs, cur_pb)}
    except Exception:
        return None


def _yoy_series(myd, key):
    """从 multi_year_data 取同比序列; 源无 yoy 键时相邻年自算兜底。返回 [小数|None]"""
    years = sorted(myd.keys())
    direct = [myd[y].get(f"{key}_yoy") if key != "net_profit" else myd[y].get("profit_yoy")
              for y in years]
    if all(v is not None for v in direct) and any(direct):
        return direct
    # 自算兜底
    out = []
    for i, y in enumerate(years):
        v_prev = myd[years[i - 1]].get(key) if i > 0 else None
        v_cur = myd[y].get(key)
        if i > 0 and v_cur is not None and v_prev not in (None, 0) and (v_prev > 0) and (v_cur > 0):
            out.append(round(v_cur / v_prev - 1, 4))
        else:
            out.append(None)
    return out


def _arrow(vals, pct_mode=False):
    """趋势箭头: 末值 vs 前值。pct_mode 用于同比行(加速/减速)。"""
    v = [x for x in vals if x is not None]
    if len(v) < 2:
        return ""
    a, b = float(v[-2]), float(v[-1])
    if pct_mode:
        return "↑加速" if b > a else ("↓减速" if b < a else "→")
    if b > a * 1.05:
        return "↑"
    if b < a * 0.95:
        return "↓"
    return "→"


def _fund_report_sections(s: dict):
    """v5.3.3(G): 按 stock-analysis 标准组装基本面三段。

    Returns (trend_table_md, review_table_md, core_judgment, valuation_md)
    全程 .get() 容错——任何数据缺失只降级展示, 不抛异常。
    """
    myd = s.get('multi_year_data') or {}
    fd = s.get('fund_data') or {}
    prof = fd.get('profitability') or {}
    growth = fd.get('growth') or {}
    health = fd.get('health') or {}
    years = sorted(myd.keys())

    # ── 段1: 5年财务趋势表（含同比/扣非行 + 趋势列）──
    trend_md = ""
    if myd and len(myd) >= 2:
        rev_series = [myd[y].get('revenue') for y in years]
        np_series = [myd[y].get('net_profit') for y in years]
        rev_yoy = _yoy_series(myd, 'revenue')
        np_yoy = _yoy_series(myd, 'net_profit')
        ded_series = [myd[y].get('deducted_profit') for y in years]
        has_ded = any(v is not None for v in ded_series)

        def _fmt_row(label, fmt_vals, arrow):
            return f"| {label} | " + " | ".join(fmt_vals + [arrow]) + " |"

        rows = []
        rows.append(_fmt_row("营收(亿)",
                             [f"{v/1e8:.1f}" if v is not None else "—" for v in rev_series],
                             _arrow(rev_series)))
        rows.append(_fmt_row("营收同比",
                             [(f"+{v*100:.1f}%" if v is not None else "—") for v in rev_yoy],
                             _arrow(rev_yoy, pct_mode=True)))
        rows.append(_fmt_row("净利(亿)",
                             [f"{v/1e8:.1f}" if v is not None else "—" for v in np_series],
                             _arrow(np_series)))
        rows.append(_fmt_row("净利同比",
                             [(f"{'+' if v is not None and v > 0 else ''}{v*100:.1f}%" if v is not None else "—")
                              for v in np_yoy],
                             _arrow(np_yoy, pct_mode=True)))
        if has_ded:
            rows.append(_fmt_row("扣非净利(亿)",
                                 [f"{v/1e8:.1f}" if v is not None else "—" for v in ded_series],
                                 _arrow(ded_series)))
            ded_yoy = [myd[y].get('deducted_yoy') for y in years]
            if any(v is not None for v in ded_yoy):
                rows.append(_fmt_row("扣非同比",
                                     [(f"{'+' if v is not None and v > 0 else ''}{v*100:.1f}%" if v is not None else "—")
                                      for v in ded_yoy],
                                     _arrow(ded_yoy, pct_mode=True)))
        gp_series = [myd[y].get('gp_margin') for y in years]
        roe_series = [myd[y].get('roe') for y in years]
        liab_series = [myd[y].get('liability') for y in years]
        rows.append(_fmt_row("毛利率",
                             [f"{v*100:.1f}%" if v is not None else "—" for v in gp_series],
                             _arrow(gp_series)))
        rows.append(_fmt_row("ROE",
                             [f"{v*100:.1f}%" if v is not None else "—" for v in roe_series],
                             _arrow(roe_series)))
        rows.append(_fmt_row("资产负债率",
                             [f"{v*100:.1f}%" if v is not None else "—" for v in liab_series],
                             _arrow(liab_series)))

        header = "| 指标 | " + " | ".join(years) + " | 趋势 |"
        sep = "|:-----|" + ":---:|" * (len(years) + 1)
        trend_md = "\n".join([header, sep] + rows)

    # ── 段2: 最新季报点评表 ──
    review_rows = []  # (指标, 数值str, 判定emoji)

    def _g(v):
        return "🟢" if v else "🔴"

    ry = growth.get('YOYRevenue')
    if ry is not None:
        review_rows.append(("营收增速(最新期)", f"{ry*100:+.1f}%",
                            "🟢" if ry > 0.20 else ("🟡" if ry > 0 else "🔴")))
    ny = growth.get('YOYNI')
    if ny is not None:
        review_rows.append(("净利增速(最新期)", f"{ny*100:+.1f}%",
                            "🟢" if ny > 0.20 else ("🟡" if ny > 0 else "🔴")))
    pq = prof.get('profitQuality')
    if pq is not None:
        review_rows.append(("利润质量 扣非/归母", f"{pq*100:.0f}%",
                            "🟢" if pq >= 0.8 else ("🟡" if pq >= 0.5 else "🔴")))
    gpm = prof.get('gpMargin')
    if gpm is not None and myd:
        gp_hist = [myd[y].get('gp_margin') for y in years if myd[y].get('gp_margin')]
        if gp_hist:
            import statistics as _stat
            mean_gp = _stat.mean(gp_hist)
            diff = (gpm - mean_gp) * 100
            review_rows.append((f"毛利率 vs 5年均值({mean_gp*100:.1f}%)", f"{diff:+.1f}pct",
                                "🟢" if diff > 2 else ("🟡" if diff > -2 else "🔴")))
    ann_roe = s.get('annual_roe')
    if ann_roe is not None:
        review_rows.append(("ROE(年报口径)", f"{ann_roe*100:.1f}%",
                            "🟢" if ann_roe > 0.15 else ("🟡" if ann_roe > 0.10 else "🔴")))
    cfnp = health.get('CFOToNP')
    if cfnp is not None:
        review_rows.append(("现金流含金量 CFO/净利", f"{cfnp*100:.0f}%",
                            "🟢" if cfnp >= 1.0 else ("🟡" if cfnp >= 0.5 else "🔴")))
    cr = health.get('currentRatio')
    if cr is not None:
        review_rows.append(("流动比率", f"{cr:.2f}",
                            "🟢" if cr >= 2.0 else ("🟡" if cr >= 1.0 else "🔴")))
    inv_d = health.get('inventory_days')
    ar_d = health.get('receivable_days')
    if inv_d is not None:
        prev_inv = myd.get(str(int(s.get('data_date', '0')[:4]) - 1), {}).get('inventory_days') \
            if s.get('data_date') else None
        note = ""
        if prev_inv:
            d = inv_d - prev_inv
            note = f"({'较年报改善' if d < 0 else '较年报走阔'}{abs(d):.0f}天)"
        review_rows.append(("存货周转天数", f"{inv_d:.0f}天{note}", "🟡"))
    if ar_d is not None:
        review_rows.append(("应收周转天数", f"{ar_d:.0f}天", "🟡"))

    review_md = ""
    core_judgment = ""
    if review_rows:
        lines = ["| 指标 | 数值 | 判定 |", "|:-----|:-----|:---:|"]
        lines += [f"| {n} | {v} | {j} |" for n, v, j in review_rows]
        review_md = "\n".join(lines)
        n_green = sum(1 for r in review_rows if r[2] == "🟢")
        reds = [r for r in review_rows if r[2] == "🔴"]
        if reds:
            core_judgment = "整体健康但需跟踪：" + "、".join(f"{n}({v})" for n, v, _ in reds[:3])
        elif n_green >= max(3, len(review_rows) - 2):
            core_judgment = "多维度互证，财务质量健康"
        else:
            core_judgment = "指标中性，无突出风险亦无突出亮点"

    # ── 段3: 估值历史分位 ──
    val_md = ""
    vp = _valuation_percentile(s.get('code'), s.get('pe'), s.get('pb'))
    if vp:
        def _vline(name, stt):
            if not stt:
                return None
            pct_s = f"**{stt['pct']:.1f}%**" if 'pct' in stt else "—"
            cur_s = f"{stt['cur']}" if 'cur' in stt else "—"
            return (f"| {name} | {cur_s} | {stt['min']} ~ {stt['max']}"
                    f"（中位{stt['median']}） | {pct_s} |")
        l1, l2 = _vline("PE-TTM", vp.get('pe')), _vline("PB", vp.get('pb'))
        if l1 or l2:
            val_md = "\n".join(["| 指标 | 当前 | 7年区间 | 历史分位 |",
                                "|:-----|:-----|:--------|:-------:|"]
                               + [x for x in (l1, l2) if x])

    return trend_md, review_md, core_judgment, val_md


def generate_md_report(s: dict, output_dir: str):
    """生成四维评分 Markdown 报告（含5年趋势表）"""
    md_path = os.path.join(output_dir, f"{s['code']}_score_report.md")

    # v5.0.1 P2：数据降级标记 — 区分"评分差"和"数据缺失/降级"
    # 必须在 f-string content 之前计算（f-string 求值时需要变量已定义）
    _dd = s.get('data_degraded') or {}
    _degraded_parts = []
    if _dd.get('tech'):
        _degraded_parts.append("⚠️ 技术分降级：用扫描评分估算（数据或分析异常）")
    if _dd.get('fund'):
        _degraded_parts.append(f"⚠️ 基本面降级：{_dd.get('fund_reason', '数据缺失')}（评分=50中性）")
    _degraded_warning = "\n".join(f"- {p}" for p in _degraded_parts)

    # ── v5.3.3(G): 按 stock-analysis 标准组装基本面段（趋势/季报点评/估值分位）──
    myd = s.get('multi_year_data')
    trend_md, review_md, core_judgment, val_md = _fund_report_sections(s)

    # 趋势统计行 + CAGR（F-5 保留项, 拼在趋势表后）
    _tail_lines = []
    roe_std = s.get('roe_std')
    rev_vol = s.get('revenue_volatility')
    tc = s.get('trend_correction', 0)
    td = s.get('trend_correction_detail', '')
    if roe_std is not None:
        _tail_lines.append(f"| ROE标准差 | {roe_std*100:.1f}% |")
    if rev_vol is not None:
        _tail_lines.append(f"| 营收波动率 | {rev_vol*100:.1f}% |")
    if tc != 0:
        _tail_lines.append(f"| 趋势修正 | {'+' if tc > 0 else ''}{tc}分 ({td}) |")

    trend_table = ""
    if trend_md:
        parts = ["\n### 5年财务趋势", trend_md]
        if _tail_lines:
            parts += ["", "| 项目 | 数值 |", "|------|------|"] + _tail_lines
        # v5.3.3(F-5): 5年营收CAGR
        try:
            years_c = sorted((myd or {}).keys())
            _rev_first = next(((myd[y] or {}).get("revenue") for y in years_c
                               if (myd[y] or {}).get("revenue")), None)
            _rev_last = next(((myd[y] or {}).get("revenue") for y in reversed(years_c)
                              if (myd[y] or {}).get("revenue")), None)
            _n_years = len(years_c) - 1
            if (_rev_first and _rev_last and _rev_first > 0 and _n_years > 0
                    and len([y for y in years_c if (myd[y] or {}).get("revenue")]) >= 3):
                import math as _math
                _cagr = (_rev_last / _rev_first) ** (1 / _n_years) - 1
                parts.append(f"\n> **{years_c[0]}-{years_c[-1]} 营收CAGR: {_cagr*100:.1f}%**"
                             f"（{_rev_first/1e8:.1f}亿 → {_rev_last/1e8:.1f}亿）")
        except Exception:
            pass
        trend_table = "\n".join(parts) + "\n"

    review_block = ""
    if review_md:
        review_block = "\n### 最新季报点评\n" + review_md
        if core_judgment:
            review_block += f"\n\n**核心判断**：{core_judgment}"
        review_block += "\n"

    valuation_block = ""
    if val_md:
        valuation_block = "\n### 估值（Baostock 7年分位）\n" + val_md + "\n"

    # v5.3.3(F-1): ROE 展示——年报为评分基准时主显年报值, 中报快照作注
    _snap_roe = s.get('roe')
    _ann_roe = s.get('annual_roe')
    if s.get('roe_basis') == 'annual' and _ann_roe is not None:
        _roe_display = f"{_ann_roe*100:.1f}%（2025年报, 评分基准"
        if _snap_roe is not None and abs(_snap_roe - _ann_roe) > 0.005:
            _yr = max((s.get('multi_year_data') or {}).keys(), default='年报')
            _roe_display += f"; 最新报告期未年化{_snap_roe*100:.1f}%"
        _roe_display += "）"
    else:
        _roe_display = f"{(_snap_roe or 0)*100:.1f}%"

    content = f"""# {s['name']}({s['code']}) 五维系统评分报告

**生成日期**: {datetime.now().strftime('%Y-%m-%d %H:%M')}
**行业**: {s.get('industry', '未知')}
**现价**: ¥{s['price']}

---

## 📊 综合评分: {s['composite']} ({s['grade']}级)

| 维度 | 得分 | 权重 | 加权 |
|------|:----:|:----:|:----:|
| 技术面 | {s['tech_score']} | {W_TECH*100:.0f}% | {s['tech_score']*W_TECH:.1f} |
| 基本面 | {s['fund_score']} | {W_FUND*100:.0f}% | {s['fund_score']*W_FUND:.1f} |
| Alpha因子 | {s.get('alpha_score', 50):.1f} | {W_ALPHA*100:.0f}% | {s.get('alpha_score', 50)*W_ALPHA:.1f} |
| 消息面 | {s['news_score']} | {W_NEWS*100:.0f}% | {s['news_score']*W_NEWS:.1f} |
| 资金面 | {s.get('fund_factor_score', '—')} | 10% | {s.get('fund_factor_score', 0)*0.10:.1f} |
| **综合** | **{s['composite']}** | 100% | — |

> 等级: A≥70 B≥60 C≥50 D<50 | {'共振惩罚已应用' if s.get('resonance') else '无共振惩罚'}

---

## 💰 仓位建议: {s['position_pct']}

{s['reason']}

---

## 🔍 技术面详情

| 项目 | 内容 |
|------|------|
| 扫描模式 | {s['pattern']} |
| 买点类型 | {s.get('buy_type', '无')} |
| 买点日期 | {s.get('buy_date', '无')} |
| 买点价格 | ¥{s.get('buy_price', 0)} |
| 最近中枢 | ZG=¥{s['zg']}, ZD=¥{s['zd']} |
| 技术得分 | {s['tech_score']} |
| 笔数 | {s['total_bis']} |
| 中枢数 | {s['total_zs']} |
| 历史买点 | {s.get('buy_count', 0)}个 |
| 历史卖点 | {s.get('sell_count', 0)}个 |

---

## 📈 基本面详情

### 当前快照
| 指标 | 数值 |
|------|------|
| ROE | {_roe_display} |
| 净利率 | {(str(round(s['np_margin']*100, 1)) + '%') if s.get('np_margin') is not None else '—'} |
| PE(TTM) | {s.get('pe') if s.get('pe') is not None else '—'}{'（获取失败）' if s.get('valuation_degraded') and s.get('pe') is None else ''} |
| PB(MRQ) | {s.get('pb') if s.get('pb') is not None else '—'}{'（获取失败）' if s.get('valuation_degraded') and s.get('pb') is None else ''} |
| 基本面总得分 | {s['fund_score']}/100 |
| 股票类型 | {classify_by_industry(s.get('industry', ''))} |
{('> ⚠️ ' + s['margin_note']) if s.get('margin_note') else ''}
{trend_table}
{review_block}
{valuation_block}
---

## 📰 消息面

- 消息得分: {s['news_score']}/100
- 数据源: {s.get('news_detail', 'N/A').split(chr(10))[0] if s.get('news_detail') else 'N/A'}

### 消息明细

{chr(10).join(s.get('news_detail', '').split(chr(10))[1:]) if s.get('news_detail') and chr(10) in s.get('news_detail', '') else '无消息明细'}

---

## 🏦 资金面

- 资金面得分: {s.get('fund_factor_score', 'N/A')}/100
- 评估方法: 筹码集中40% + 两融趋势30% + 资金流120日30%

### 资金面明细

- 筹码集中得分: {s.get('holder_score', '—') if s.get('holder_score') is not None else '—'}
- 两融趋势得分: {s.get('margin_score', '—') if s.get('margin_score') is not None else '—'}
- 资金流120日得分: {s.get('flow_score', '—') if s.get('flow_score') is not None else '—'}
- 数据来源: {s.get('fund_factor_detail', '非 Top30 候选，未补扫资金面因子')}

---

## ⚠️ 风险提示

- 本报告由五维分析系统自动生成，仅供参考，不构成投资建议
- 技术面基于日线缠论分析，未做30分钟多级别确认
- 消息面为自动化搜索摘要，可能遗漏重要信息
- 投资决策请结合个人风险承受能力和市场整体环境
{_degraded_warning}
"""
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(content)
    return md_path


def generate_summary_md(scored: list) -> str:
    """生成 MD 格式汇总表"""
    path = os.path.join(OUTPUT_BASE, f"扫描汇总_{datetime.now().strftime('%Y-%m-%d')}.md")

    # v5.3.3(E-2): 观察型信号不入推荐列表, 单独归入观察区小节
    _watch = [s for s in scored if s.get('observational')
              and s['composite'] >= COMPOSITE_THRESHOLD][:5]
    # v5.4(C-05): 推荐口径与 generate_reports 完全对齐(阈值+can_buy+非观察)——
    # 旧实现按纯分数截TopN, 低于阈值/禁建仓股被列为"推荐"且其详情链接指向
    # 不存在的个股报告文件夹(死链)。空池兜底三级降级与 generate_reports 同款。
    _qualified = [s for s in scored
                  if s['composite'] >= COMPOSITE_THRESHOLD and not s.get('observational')
                  and s.get('can_buy', True)]
    top = _qualified[:TOP_N_REPORT]
    if not top:
        top = [s for s in scored[:min(10, len(scored))]
               if not s.get('observational') and s.get('can_buy', True)] or \
              [s for s in scored[:min(10, len(scored))] if not s.get('observational')] or \
              scored[:min(10, len(scored))]
    # v2026-08-28(A2否决可见性): 全池去向核算——veto 股(composite=0/grade=D)此前
    # 从报告静默消失(2026-08-28 川投能源实锤: 33 候选里 1 只否决股无区块无原因),
    # 用户无法知道系统何时挡过雷/误杀过票。头部行给全池去向账, 观察区后加否决区块。
    _watch_all = [s for s in scored if s.get('observational')]
    _vetoed = [s for s in scored
               if '风控否决' in (s.get('reason') or '')
               or (s.get('composite', 0) == 0 and s.get('grade') == 'D')]
    _below = max(len(scored) - len(_qualified) - len(_watch_all) - len(_vetoed), 0)
    lines = [
        f"# A500 股票池智能筛选汇总",
        f"**扫描日期**: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"**候选股**: {len(scored)} 只 | **推荐**: {len(_qualified)} 只"
        f"(综合≥{COMPOSITE_THRESHOLD} 且可建仓) | 下列表展示 Top {len(top)}",
        f"**全池去向**: 推荐 {len(_qualified)} + 观察区 {len(_watch_all)}"
        f" + 风控否决 {len(_vetoed)} + 低于阈值未展示 {_below} = {len(scored)} 只",
        "",
        "## 🏆 推荐列表（按综合分降序）",
        "",
        "| # | 代码 | 名称 | 综合 | 等级 | 技术 | 基本面 | Alpha | 消息 | 资金面 | 仓位 | 模式 | 行业 |",
        "|---|------|------|:----:|:----:|:----:|:------:|:----:|:----:|:----:|:----:|------|------|",
    ]

    for i, s in enumerate(top, 1):
        icon = "🟢" if s['grade'] == 'A' else ("🟡" if s['grade'] == 'B' else ("🟠" if s['grade'] == 'C' else "🔴"))
        # v5.3.2(D-5/X1): 高综合评级但零仓位 = 技术面未过建仓线, 标⚠消解
        # "B级却0%仓位"的语义冲突(评级来自五维composite, 仓位闸门仅看tech)
        _gate_warn = (s['grade'] in ('A', 'B') and not s.get('can_buy', True))
        grade_disp = f"{icon}{s['grade']}{'⚠' if _gate_warn else ''}"
        lines.append(
            f"| {i} | {s['code']} | {s['name']} | **{s['composite']}** | "
            f"{grade_disp} | {s['tech_score']} | {s['fund_score']} | "
            f"{s.get('alpha_score', 50):.1f} | {s['news_score']} | "
            f"{s.get('fund_factor_score', '—')} | {s['position_pct']} | {s['pattern'][:20]} | "
            f"{s.get('industry', '')[:8]} |"
        )

    _n_gate = sum(1 for s in top if s['grade'] in ('A', 'B') and not s.get('can_buy', True))
    if _n_gate:
        lines.append("")
        lines.append(f"> ⚠ ×{_n_gate}: 综合评级达 B 及以上, 但技术分未过建仓线({TECH_BUY_THRESHOLD}分) → 仓位 0%。"
                     f"评级反映五维均衡度, 建仓闸门单独由技术面把守——二者口径不同属设计行为, 非数据错误。")

    # v5.3.3(E-1): 卖出冲突说明（若仍出现在全量数据中）
    _n_conflict = sum(1 for s in scored if s.get('sell_conflict'))
    if _n_conflict:
        lines.append("")
        lines.append(f"> 🚫 ×{_n_conflict}: 近{SELL_SIGNAL_SUPPRESS_DAYS}日内出现一卖/二卖 → 买卖冲突仲裁, 全部买入信号压制(v5.3.3 E-1)。")

    # v5.3.3(E-2): 观察区——几何形态信号, 可跟踪不入推荐
    if _watch:
        lines += [
            "",
            "## 👁 观察区（几何形态信号, 非缠论确认买点）",
            "",
            "> 以下为「三买形成中/突破延续」类纯几何观察信号：可入池跟踪, 但不构成买点确认, 不进推荐、仓位封顶轻仓。",
            "",
            "| 代码 | 名称 | 综合 | 等级 | 模式 | 说明 |",
            "|------|------|:----:|:----:|------|------|",
        ]
        for s in _watch:
            lines.append(
                f"| {s['code']} | {s['name']} | {s['composite']} | {s['grade']} | "
                f"{s['pattern'][:24]} | 待缠论结构确认后再评估 |"
            )

    # v2026-08-28(A2否决可见性): 风控否决区块——veto 股 composite=0 天然低于
    # 推荐阈值, 但"从报告中消失"与"如实标注"纪律相悖; 且否决可能来自关键词
    # 误杀(如制度类公告模板条文), 必须把原因交给用户人工复核。
    if _vetoed:
        lines += [
            "",
            "## ⛔ 本轮风控否决",
            "",
            "> 以下股票触发一票否决(composite=0, 仓位0), 不参与排名。否决原因来自",
            "> 消息面关键词条目级匹配, 若为模板条文误杀(如治理制度类公告引用的",
            "> 「涉嫌违法」条款文字), 请结合个股公告人工复核后放行。",
            "",
            "| 代码 | 名称 | 否决原因 | 否决前五维原值(技术/基本/消息) |",
            "|------|------|------|------|",
        ]
        for s in _vetoed:
            _vr = (s.get('reason') or '未知').replace('⛔ 风控否决: ', '').replace('|', '\\|')
            lines.append(
                f"| {s['code']} | {s['name']} | {_vr} | "
                f"{s.get('tech_score', '—')} / {s.get('fund_score', '—')} / {s.get('news_score', '—')} |"
            )

    lines += [
        "",
        "---",
        "",
        "## 📁 详细报告",
        "",
    ]
    for s in top:
        lines.append(f"- [{s['name']}({s['code']})]({s['name']}_{s['code']}/{s['code']}_score_report.md) — HTML技术分析 + 评分报告")

    lines += [
        "",
        "---",
        "",
        "## 📰 消息面摘要",
        "",
    ]
    for i, s in enumerate(top, 1):
        news_detail = s.get('news_detail', '')
        # v5.4.1(AUD-B-04): news_scan_failed 消费接线——补扫失败股在汇总区
        # 显式警示(此前标志只写不读), 防止"缺消息面"与"已检查无负面"混淆
        if s.get('news_scan_failed'):
            lines.append(f"### {i}. {s['name']}({s['code']}) — ⚠️ 消息面补扫失败"
                         f"(news_score={s.get('news_score', '—')}为占位值，需人工复核)")
            lines.append("")
            continue
        if not news_detail:
            continue
        # 取第一行（数据源汇总）+ 消息明细
        detail_lines = news_detail.split(chr(10))
        summary_line = detail_lines[0] if detail_lines else ''
        msg_lines = detail_lines[1:] if len(detail_lines) > 1 else []
        # 只取前5条消息，避免太长
        msg_preview = msg_lines[:5]
        lines.append(f"### {i}. {s['name']}({s['code']}) — 消息分 {s['news_score']}")
        lines.append(f"  {summary_line}")
        if msg_preview:
            lines.append("  关键消息:")
            for ml in msg_preview:
                lines.append(f"  - {ml}")
        if len(msg_lines) > 5:
            lines.append(f"  ... 共 {len(msg_lines)} 条消息，详见个股报告")
        lines.append("")

    lines += [
        "",
        "---",
        "*报告由 Hermes 五维分析系统自动生成*",
    ]

    content = "\n".join(lines)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)
    return path


def generate_summary_excel(scored: list) -> str:
    """生成 Excel 格式汇总表"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [跳过] openpyxl 未安装，不生成 Excel")
        return ""

    path = os.path.join(OUTPUT_BASE, f"扫描汇总_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    # v5.4.1(P3): sheet 名与实装一致——旧名"Top10"而实际写入全部推荐股
    ws.title = "推荐汇总"

    # 表头
    headers = ['排名', '代码', '名称', '综合分', '等级', '技术分', '基本面分', 'Alpha分', '消息分', '资金面分',
               '仓位', '现价', '模式', '买点类型', '买点日期', '买点价', 'ZG', 'ZD',
               'ROE%', 'PE', 'PB', '行业', '共振惩罚', '可建仓', '备注']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    grade_colors = {
        'A': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'B': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'C': PatternFill(start_color='F4B4C2', end_color='F4B4C2', fill_type='solid'),
        'D': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
    }

    # v5.4(C-05): 与 generate_reports / 汇总MD 口径对齐(阈值+can_buy+非观察)
    top = [s for s in scored
           if s['composite'] >= COMPOSITE_THRESHOLD and not s.get('observational')
           and s.get('can_buy', True)][:TOP_N_REPORT]
    if not top:
        top = [s for s in scored[:min(10, len(scored))]
               if not s.get('observational') and s.get('can_buy', True)] or \
              [s for s in scored[:min(10, len(scored))] if not s.get('observational')] or \
              scored[:min(10, len(scored))]
    for row_idx, s in enumerate(top, 2):
        values = [
            row_idx - 1, s['code'], s['name'], s['composite'], s['grade'],
            s['tech_score'], s['fund_score'], s.get('alpha_score', 50), s['news_score'],
            s.get('fund_factor_score', '—'),
            s['position_pct'], s['price'], s['pattern'][:30],
            s.get('buy_type', ''), s.get('buy_date', ''), s.get('buy_price', 0),
            s['zg'], s['zd'], s.get('roe', 0), s.get('pe', 0), s.get('pb', 0),
            s.get('industry', ''), '是' if s.get('resonance') else '否',
            '是' if s.get('can_buy') else '否',
            '观察' if s.get('observational') else ('卖出冲突' if s.get('sell_conflict') else ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 等级着色
        grade_fill = grade_colors.get(s['grade'])
        if grade_fill:
            ws.cell(row=row_idx, column=2).fill = grade_fill  # 代码
            ws.cell(row=row_idx, column=5).fill = grade_fill  # 等级

    # 列宽
    col_widths = [5, 8, 10, 8, 5, 8, 9, 8, 8, 8, 6, 9, 25, 8, 14, 9, 9, 9, 8, 8, 8, 10, 8, 6, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = 'A2'

    # Sheet 2: 全部候选
    ws2 = wb.create_sheet("全部候选")
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, s in enumerate(scored, 2):
        values = [
            row_idx - 1, s['code'], s['name'], s['composite'], s['grade'],
            s['tech_score'], s['fund_score'], s.get('alpha_score', 50), s['news_score'],
            s.get('fund_factor_score', '—'),
            s['position_pct'], s['price'], s['pattern'][:30],
            s.get('buy_type', ''), s.get('buy_date', ''), s.get('buy_price', 0),
            s['zg'], s['zd'], s.get('roe', 0), s.get('pe', 0), s.get('pb', 0),
            s.get('industry', ''), '是' if s.get('resonance') else '否',
            '是' if s.get('can_buy') else '否',
            '观察' if s.get('observational') else ('卖出冲突' if s.get('sell_conflict') else ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    try:
        wb.save(path)
    except PermissionError:
        # 文件被占用（如用户 Excel 打开中），加时间戳保存
        alt_path = os.path.join(OUTPUT_BASE, f"扫描汇总_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx")
        wb.save(alt_path)
        path = alt_path
    return path


def generate_reports(scored: list, gen_summary: bool = True):
    """为每只推荐股生成 HTML + MD 报告

    gen_summary=False (v2026-08-28 编排优化): 跳过 MD/Excel 汇总表生成——
    Phase 2+3 阶段的三维口径汇总(α=50中性)注定被 Step4 五维重算覆盖,
    提前生成只会造成"同一份汇总表被反复覆盖+旧文件夹被反复清理"的三遍劳动。
    编排方(pipeline)在 Phase 2+3 传 False, 汇总表由 fund_factor_rescore
    --report 五维终版一次性生成。手动分步跑不传参则行为不变。"""
    os.makedirs(OUTPUT_BASE, exist_ok=True)

    # v5.3.1(F13): 财务风控降级可见性——AKShare 失败的股票资不抵债/连亏
    # 检查被跳过, 报告生成时必须让用户知道风控覆盖缺口
    try:
        from risk_filter import get_risk_check_degraded
        _degraded = get_risk_check_degraded()
        if _degraded:
            print(f"\n⚠⚠ 风控降级警告: {len(_degraded)} 只股票财务数据获取失败, "
                  f"资不抵债/连亏检查被跳过: {', '.join(_degraded[:10])}"
                  f"{'...' if len(_degraded) > 10 else ''}")
            _top_degraded = [s['code'] for s in scored[:20] if s.get('code') in set(_degraded)]
            if _top_degraded:
                print(f"⚠⚠ 其中进入 Top20: {', '.join(_top_degraded)} —— 请人工核查其财务风险!")
    except Exception:
        pass

    # v5.3.3(E-2): 观察型几何信号不进推荐列表——单独归入观察区
    # 终审A3(2026-08-23): 补排 can_buy=False——E-1 卖出冲突/veto/severe 压制股
    # 以0%仓位现身🏆推荐列表与"买入信号全部压制"语义矛盾(composite_scorer:352
    # 对 recent_top_sell 置 can_buy=False)。弱市兜底同步；仅当全池皆压制时
    # 才按原样展示并在🚫脚注/'卖出冲突'列披露(保持报告非空的最后防线)。
    _watch = [s for s in scored if s.get('observational')
              and s['composite'] >= COMPOSITE_THRESHOLD][:TOP_N_REPORT]
    top = [s for s in scored
           if s['composite'] >= COMPOSITE_THRESHOLD and not s.get('observational')
           and s.get('can_buy', True)][:TOP_N_REPORT]
    if not top:
        top = [s for s in scored[:min(10, len(scored))]
               if not s.get('observational') and s.get('can_buy', True)] or \
              [s for s in scored[:min(10, len(scored))] if not s.get('observational')] or \
              scored[:min(10, len(scored))]
    if _watch:
        print(f"\n[E-2] {len(_watch)} 只观察型信号(三买形成中/突破延续)不入推荐, 移入观察区: "
              f"{', '.join(s['code'] + ' ' + s.get('name', '') for s in _watch)}")

    print(f"\n[Phase 3] 生成 {len(top)} 只股票的报告...")

    # v5.0.1 修复：清理不在推荐列表中的旧文件夹
    # 五维重算后部分股票跌破阈值（如四维62→五维59），旧文件夹会残留。
    # 生成前先删除不在 top 列表中的 `*_代码` 文件夹，保持输出目录与推荐一致。
    try:
        _top_codes = {s['code'] for s in top}
        for _item in os.listdir(OUTPUT_BASE):
            _dir = os.path.join(OUTPUT_BASE, _item)
            if not os.path.isdir(_dir):
                continue
            _parts = _item.rsplit('_', 1)
            if len(_parts) == 2 and _parts[1].isdigit() and len(_parts[1]) == 6:
                if _parts[1] not in _top_codes:
                    import shutil
                    shutil.rmtree(_dir, ignore_errors=True)
                    print(f"  清理旧文件夹: {_item}")
    except Exception as _e:
        print(f"  [警告] 旧文件夹清理失败: {_e}")

    for i, s in enumerate(top):
        stock_dir = os.path.join(OUTPUT_BASE, f"{s['name']}_{s['code']}")
        os.makedirs(stock_dir, exist_ok=True)

        print(f"  [{i+1}/{len(top)}] {s['name']}({s['code']}) ...", end=" ", flush=True)
        html_ok = generate_html_report(s, stock_dir)
        md_ok = generate_md_report(s, stock_dir)
        print(f"HTML={'✓' if html_ok else '✗'} MD={'✓' if md_ok else '✗'}")

    # 生成汇总表（gen_summary=False 时跳过——三维口径中间产物, 见函数 docstring）
    if gen_summary:
        md_summary = generate_summary_md(scored)
        xlsx_summary = generate_summary_excel(scored)

        print(f"\n[Phase 3] 完成!")
        print(f"  个股报告目录: {OUTPUT_BASE}/")
        print(f"  MD总表: {md_summary}")
        if xlsx_summary:
            print(f"  Excel总表: {xlsx_summary}")
    else:
        print(f"\n[Phase 3] 完成! (跳过汇总表——由五维重算阶段统一生成)")
        print(f"  个股报告目录: {OUTPUT_BASE}/")


# ============================================================
# 主流程
# ============================================================

def main():
    args = sys.argv[1:]
    phase1_only = '--phase1-only' in args
    from_cache = '--from-cache' in args
    test_n = 0

    for i, a in enumerate(args):
        if a == '--test' and i + 1 < len(args):
            test_n = int(args[i + 1])
    # v2026-08-28 编排优化: pipeline 的 Phase 2+3 传 --skip-summary,
    # 三维口径汇总表(α=50中性)不再生成, 由五维重算终版一次性输出
    skip_summary = '--skip-summary' in args

    t0 = time.time()

    # Phase 1
    if from_cache:
        candidates, cache_data = load_phase1_cache()
    else:
        candidates, cache_data = run_phase1(test_n)

    if phase1_only:
        print("\n仅 Phase 1，完成。")
        return

    if not candidates:
        print("\n无候选股，退出。")
        return

    # Phase 2
    print(f"\n{'='*60}")
    print(f"Phase 2: 三维深度评估 ({len(candidates)} 只, 分{BATCH_COUNT}批, 批次间暂停{BATCH_PAUSE}s)")
    print(f"{'='*60}")
    scored = run_phase2(candidates, batch_count=BATCH_COUNT, batch_pause=BATCH_PAUSE)

    if not scored:
        print("\n无有效评分，退出。")
        return

    # Phase 3
    print(f"\n{'='*60}")
    print(f"Phase 3: 报告生成")
    print(f"{'='*60}")
    generate_reports(scored, gen_summary=not skip_summary)

    elapsed = time.time() - t0
    print(f"\n{'='*60}")
    print(f"全流程完成! 总耗时 {elapsed:.0f}s ({elapsed/60:.1f}min)")
    print(f"候选股: {len(candidates)} → 推荐: {len([s for s in scored if s['composite'] >= COMPOSITE_THRESHOLD])} 只")
    print(f"输出目录: {OUTPUT_BASE}/")

    # 写入成功标记（供 a500_backtest.py 检查依赖）
    from cron_utils import FlagSignals
    flag_date = datetime.now().strftime("%Y-%m-%d")
    flag_path = FlagSignals.write("a500_scan_done", flag_date,
                                   extra={"candidates": len(candidates),
                                          "scored": len(scored),
                                          "elapsed_s": round(elapsed, 1)})
    print(f"  成功标记: {flag_path}")


if __name__ == "__main__":
    main()
