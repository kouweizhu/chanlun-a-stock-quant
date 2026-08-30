"""excel_report.py — 缠论分析Excel报告生成器

配合 report_generator.py 使用，每次分析生成一个.xlsx文件，
包含：
  1. 交易指令单（当前动作、仓位、止损、目标、理由）
  2. 买卖点历史（全部买卖点明细、置信度）
  3. 回测统计（总收益率、胜率、夏普等）
  4. 关键结构（中枢、笔、当前价位）

依赖：openpyxl
"""

import os
from date_utils import date_to_str, parse_date_to_datetime
from datetime import datetime
from typing import Optional, Dict, List
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ======================== 样式常量 ========================
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL_BUY = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
HEADER_FILL_SELL = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
HEADER_FILL_INFO = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FILL_ACCENT = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")

ROW_FILL_ALT = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
ROW_FILL_BUY = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid")
ROW_FILL_SELL = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

BUY_FONT = Font(color="27AE60", bold=True)
SELL_FONT = Font(color="E74C3C", bold=True)
HOLD_FONT = Font(color="7F8C8D", bold=True)

TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="2C3E50")
LABEL_FONT = Font(name="微软雅黑", bold=True, size=10, color="34495E")
VALUE_FONT = Font(name="微软雅黑", size=10)
SECTION_FONT = Font(name="微软雅黑", bold=True, size=12, color="2980B9")

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color="2C3E50")
)

WARN_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")


def _style_header_row(ws, row: int, fill: PatternFill, max_col: int):
    """给标题行加样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_cell(cell, alt: bool = False):
    """给数据单元格加基础样式"""
    cell.font = VALUE_FONT
    if alt:
        cell.fill = ROW_FILL_ALT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """自适应列宽"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            # 估算中文字符宽度（中文≈2个英文字符）
            cjk_count = sum(1 for c in val if '\u4e00' <= c <= '\u9fff')
            cell_len = len(val) + cjk_count
            if cell_len > max_len:
                max_len = cell_len
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def _write_section_title(ws, row: int, title: str, max_col: int):
    """写入分区标题"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    cell.border = BOTTOM_BORDER
    return row + 1


# ======================== 核心函数 ========================

def generate_excel_report(
    symbol: str,
    name: str,
    daily_analyzer,
    signal=None,
    backtest_stats: Optional[Dict] = None,
    reference_price: Optional[float] = None,
    output_dir: Optional[str] = None,
) -> str:
    """生成缠论分析Excel报告

    参数:
        symbol: 股票代码
        name: 股票名称
        daily_analyzer: 日线ChanLunAnalyzer实例
        signal: TradeSignal 交易信号
        backtest_stats: 回测统计字典（来自 backtest_engine.run_single）
        reference_price: 参考价（用于判断当前价格位置）
        output_dir: 输出目录（默认缠论项目目录）

    返回:
        Excel文件路径
    """
    wb = Workbook()

    # ========== Sheet 1: 交易指令单 ==========
    ws1 = wb.active
    ws1.title = "交易指令单"
    _build_signal_sheet(ws1, symbol, name, signal, daily_analyzer, reference_price, backtest_stats)

    # ========== Sheet 2: 买卖点历史 ==========
    ws2 = wb.create_sheet("买卖点历史")
    _build_points_sheet(ws2, daily_analyzer)

    # ========== Sheet 3: 回测统计 ==========
    ws3 = wb.create_sheet("回测统计")
    _build_backtest_sheet(ws3, backtest_stats, signal)

    # ========== Sheet 4: 关键结构 ==========
    ws4 = wb.create_sheet("关键结构")
    _build_structure_sheet(ws4, daily_analyzer)

    # 保存
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(__file__))

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"{symbol}_chanlun_report.xlsx")

    # 删除已有文件避免权限问题
    if os.path.exists(filepath):
        os.remove(filepath)

    wb.save(filepath)
    return filepath


# ======================== Sheet 1: 交易指令单 ========================

def _build_signal_sheet(ws, symbol: str, name: str, signal, analyzer, reference_price: Optional[float] = None, backtest_stats: Optional[Dict] = None):
    """构建交易指令单页面"""
    max_col = 6
    current_price = analyzer.klines[-1].close if analyzer.klines else 0
    latest_date = analyzer.klines[-1].date if analyzer.klines else ""

    # ---- 标题 ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1,
                         value=f"{name}({symbol}) — 缠论择时交易指令单")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---- 基本信息 ----
    row = 3
    row = _write_section_title(ws, row, " 基本信息", max_col)
    row += 1

    info_data = [
        ("股票代码", symbol),
        ("股票名称", name),
        ("分析日期", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("数据截至", latest_date),
        ("当前价格", f"¥{current_price:.2f}"),
        ("参考价", f"¥{reference_price:.2f}" if reference_price else "未指定"),
        ("K线数量", f"{len(analyzer.klines)}根"),
        ("笔数量", f"{len(analyzer.bis)}笔"),
        ("中枢数量", f"{len(analyzer.zhongshus)}个"),
        ("买卖点", f"{sum(1 for p in analyzer.buy_sell_points if p.type=='buy')}买 / {sum(1 for p in analyzer.buy_sell_points if p.type=='sell')}卖"),
    ]

    for i, (label, value) in enumerate(info_data):
        c1 = ws.cell(row=row + i, column=1, value=label)
        c1.font = LABEL_FONT
        c1.border = THIN_BORDER
        c1.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=max_col)
        c2 = ws.cell(row=row + i, column=2, value=value)
        c2.font = VALUE_FONT
        c2.border = THIN_BORDER
        c2.alignment = Alignment(vertical="center")
        if i % 2 == 1:
            c1.fill = ROW_FILL_ALT
            c2.fill = ROW_FILL_ALT

    # ---- 交易信号 ----
    row += len(info_data) + 1
    signal_start = row
    row = _write_section_title(ws, row, " 交易信号", max_col)
    row += 1

    if signal is None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        c = ws.cell(row=row, column=1, value="⚠ 未能生成交易信号")
        c.font = Font(color="E74C3C", size=10)
        c.fill = WARN_FILL
        c.border = THIN_BORDER
        _auto_width(ws)
        return

    # 信号表头
    headers = ["项目", "内容", "", "", "", ""]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_ACCENT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1

    # 动作 + 颜色
    action_emoji = {"BUY": "🟢 买入", "SELL": "🔴 卖出", "HOLD": "⚪ 持有"}
    action_fill = {"BUY": ROW_FILL_BUY, "SELL": ROW_FILL_SELL, "HOLD": ROW_FILL_ALT}
    action_font = {"BUY": BUY_FONT, "SELL": SELL_FONT, "HOLD": HOLD_FONT}

    signal_rows = [
        ("执行动作", action_emoji.get(signal.action, signal.action)),
        ("建议仓位", f"{signal.position_size * 100:.0f}%"),
        ("入场价格", f"¥{signal.entry_price:.2f}"),
        ("止损价格", f"¥{signal.stop_loss:.2f}" if signal.stop_loss > 0 else "—"),
        ("目标价格", f"¥{signal.take_profit:.2f}" if signal.take_profit > 0 else "—"),
        ("逻辑理由", signal.reason),
        ("优先级", {"HIGH": "⚠ 高（立即执行）", "MEDIUM": "● 中（观察确认）", "LOW": "○ 低（暂不操作）"}.get(signal.urgency, signal.urgency)),
    ]

    # 盈亏比（如果有）
    if signal.action != "HOLD" and signal.stop_loss > 0 and signal.take_profit > 0:
        denom = abs(signal.stop_loss - signal.entry_price)
        if denom > 0.001:
            rr_ratio = abs((signal.take_profit - signal.entry_price) / denom)
            signal_rows.append(("盈亏比", f"{rr_ratio:.2f}"))
        if signal.stop_loss > 0:
            sl_pct = (signal.stop_loss / signal.entry_price - 1) * 100
            signal_rows.append(("止损幅度", f"{sl_pct:.2f}%"))
        if signal.take_profit > 0:
            tp_pct = (signal.take_profit / signal.entry_price - 1) * 100
            signal_rows.append(("目标涨幅", f"{tp_pct:.2f}%"))

    for i, (label, value) in enumerate(signal_rows):
        alt = i % 2 == 1
        c1 = ws.cell(row=row + i, column=1, value=label)
        c1.font = LABEL_FONT
        _style_data_cell(c1, alt)
        c1.alignment = Alignment(horizontal="right", vertical="center")
        if label == "执行动作" and signal.action in action_fill:
            c1.fill = action_fill[signal.action]

        ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=max_col)
        c2 = ws.cell(row=row + i, column=2, value=value)
        _style_data_cell(c2, alt)
        if label == "执行动作" and signal.action in action_font:
            c2.font = action_font[signal.action]
        if label == "逻辑理由":
            c2.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[row + i].height = 40

    # ---- 当前状态 ----
    row += len(signal_rows) + 1
    row = _write_section_title(ws, row, " 当前状态评估", max_col)
    row += 1

    # 判断当前相对买卖点的位置
    status_lines = []
    buy_points = [p for p in analyzer.buy_sell_points if p.type == "buy"]
    sell_points = [p for p in analyzer.buy_sell_points if p.type == "sell"]

    if buy_points:
        last_buy = buy_points[-1]
        status_lines.append(("最新买点", f"{last_buy.date} ¥{last_buy.price:.2f}"))
    if sell_points:
        last_sell = sell_points[-1]
        status_lines.append(("最新卖点", f"{last_sell.date} ¥{last_sell.price:.2f}"))

    # 相对中枢位置
    if analyzer.zhongshus:
        last_zs = analyzer.zhongshus[-1]
        if current_price > last_zs.zg:
            zone = "中枢上方（突破状态）"
        elif current_price < last_zs.zd:
            zone = "中枢下方（弱势状态）"
        else:
            zone = "中枢内部（震荡状态）"
        status_lines.append(("相对中枢", f"{zone}  (¥{last_zs.zd:.2f}~¥{last_zs.zg:.2f})"))

    if analyzer.bis:
        last_bi = analyzer.bis[-1]
        bi_dir = "上升" if last_bi.direction == "up" else "下降"
        bi_range = f"¥{last_bi.start_price:.2f} → ¥{last_bi.end_price:.2f}"
        status_lines.append(("最新笔方向", f"{bi_dir}笔 {bi_range}"))

    if backtest_stats:
        status_lines.append(("回测总收益", f"{backtest_stats.get('total_return', 0):+.2f}%"))
        status_lines.append(("回测胜率", f"{backtest_stats.get('win_rate', 0):.1f}%"))

    for i, (label, value) in enumerate(status_lines):
        alt = i % 2 == 1
        c1 = ws.cell(row=row + i, column=1, value=label)
        c1.font = LABEL_FONT
        _style_data_cell(c1, alt)
        c1.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=max_col)
        c2 = ws.cell(row=row + i, column=2, value=value)
        _style_data_cell(c2, alt)

    _auto_width(ws)


# ======================== Sheet 2: 买卖点历史 ========================

def _build_points_sheet(ws, analyzer):
    """构建买卖点历史页面"""
    points = analyzer.buy_sell_points
    if not points:
        ws.cell(row=1, column=1, value="无买卖点数据").font = VALUE_FONT
        return

    headers = ["序号", "类型", "级别", "日期", "价格", "置信度", "高置信", "确认方式", "理由"]
    max_col = len(headers)

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title = ws.cell(row=1, column=1, value=f"买卖点历史明细（共{len(points)}个）")
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # 表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_INFO
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 数据
    conf_type_map = {"direct": "直接确认", "divergence": "笔背驰", "none": "无确认"}
    for i, p in enumerate(points):
        r = 4 + i
        alt = i % 2 == 1
        is_buy = p.type == "buy"

        fill = ROW_FILL_BUY if is_buy else ROW_FILL_SELL if p.type == "sell" else ROW_FILL_ALT
        type_font = BUY_FONT if is_buy else SELL_FONT

        vals = [
            i + 1,
            "买入" if is_buy else "卖出",
            f"{p.level}类",
            str(p.date),
            f"¥{p.price:.2f}",
            "",
            "",
            "",
            p.reason,
        ]

        # 多级别确认信息
        ml = getattr(p, "multilevel_confirmation", None)
        if ml:
            vals[5] = f"{ml.get('confidence_score', 0)}/5"
            vals[6] = "⭐" if ml.get("high_confidence") else ""
            ct = ml.get("confirmation_type", "none")
            vals[7] = conf_type_map.get(ct, ct)
        else:
            vals[5] = "—"
            vals[6] = ""
            vals[7] = "—"

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            _style_data_cell(cell)
            cell.fill = fill
            if col == 2:
                cell.font = type_font
            if col == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in (5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    col_widths = [6, 8, 8, 14, 10, 10, 8, 12, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ======================== Sheet 3: 回测统计 ========================

def _build_backtest_sheet(ws, stats: Optional[Dict], signal):
    """构建回测统计页面"""
    max_col = 4

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title = ws.cell(row=1, column=1, value="漏斗过滤法回测统计")
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    if not stats:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
        c = ws.cell(row=3, column=1, value="⚠ 无回测数据（尚未运行回测）")
        c.font = Font(color="E74C3C", size=10)
        c.fill = WARN_FILL
        c.border = THIN_BORDER
        _auto_width(ws)
        return

    # ---- 核心指标 ----
    row = 3
    row = _write_section_title(ws, row, " 核心指标", max_col)
    row += 1

    metrics = [
        ("总收益率", f"{stats.get('total_return', 0):+.2f}%"),
        ("年化收益率", f"{stats.get('annual_return', 0):+.2f}%"),
        ("最终资产", f"¥{stats.get('final_value', 0):,.2f}"),
        ("回测周期", f"{stats.get('years', 0):.2f}年"),
        ("初始本金", f"¥{stats.get('initial_capital', 2000000):,.0f}"),
    ]

    for i, (label, value) in enumerate(metrics):
        alt = i % 2 == 1
        c1 = ws.cell(row=row + i, column=1, value=label)
        c1.font = LABEL_FONT
        _style_data_cell(c1, alt)
        c1.alignment = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=max_col)
        c2 = ws.cell(row=row + i, column=2, value=value)
        _style_data_cell(c2, alt)
        # 总收益率着色
        if label == "总收益率":
            ret = stats.get("total_return", 0)
            if ret > 0:
                c2.font = Font(color="27AE60", bold=True, size=11)
            elif ret < 0:
                c2.font = Font(color="E74C3C", bold=True, size=11)

    # ---- 风控指标 ----
    row += len(metrics) + 1
    row = _write_section_title(ws, row, " 风控指标", max_col)
    row += 1

    risk_metrics = [
        ("最大回撤", f"{stats.get('max_drawdown', 0):.2f}%"),
        ("夏普比率", f"{stats.get('sharpe_ratio', 0):.2f}"),
        ("利润因子", f"{stats.get('profit_factor', 0):.2f}"),
    ]
    # 着色夏普
    sharpe = stats.get('sharpe_ratio', 0)
    if sharpe >= 1.5:
        risk_metrics.append(("夏普评价", "优秀（>1.5）"))
    elif sharpe >= 1.0:
        risk_metrics.append(("夏普评价", "良好（>1.0）"))
    elif sharpe >= 0.5:
        risk_metrics.append(("夏普评价", "一般（>0.5）"))
    else:
        risk_metrics.append(("夏普评价", "较差（<0.5）"))

    for i, (label, value) in enumerate(risk_metrics):
        alt = i % 2 == 1
        c1 = ws.cell(row=row + i, column=1, value=label)
        c1.font = LABEL_FONT
        _style_data_cell(c1, alt)
        c1.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=max_col)
        c2 = ws.cell(row=row + i, column=2, value=value)
        _style_data_cell(c2, alt)

    # ---- 交易统计 ----
    row += len(risk_metrics) + 1
    row = _write_section_title(ws, row, " 交易统计", max_col)
    row += 1

    trade_count = stats.get("total_trades", 0)
    trade_metrics = [
        ("总交易次数", f"{trade_count}次"),
        ("胜率", f"{stats.get('win_rate', 0):.1f}%"),
        ("平均盈利", f"{stats.get('avg_win', 0):+.2f}%"),
        ("平均亏损", f"{stats.get('avg_loss', 0):+.2f}%"),
        ("盈亏比", f"{stats.get('profit_loss_ratio', 0):.2f}"),
    ]

    for i, (label, value) in enumerate(trade_metrics):
        alt = i % 2 == 1
        c1 = ws.cell(row=row + i, column=1, value=label)
        c1.font = LABEL_FONT
        _style_data_cell(c1, alt)
        c1.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=max_col)
        c2 = ws.cell(row=row + i, column=2, value=value)
        _style_data_cell(c2, alt)

    # ---- 交易明细 ----
    row += len(trade_metrics) + 1
    trades_list = stats.get('trades', [])
    if trades_list:
        row = _write_section_title(ws, row, f" 交易明细（共{len(trades_list)}笔）", max_col)
        row += 1

        trade_headers = ["买入日", "卖出日", "方向", "买入价", "卖出价", "盈亏%", "股数", "持仓天", "买入理由", "卖出理由"]
        for col, h in enumerate(trade_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        row += 1

        for i, t in enumerate(trades_list):
            alt = i % 2 == 1
            pnl = t.get('pnl_pct', 0)
            vals = [
                str(t.get('entry_date', '')),
                str(t.get('exit_date', '')),
                "多头",
                f"¥{t.get('entry_price', 0):.2f}",
                f"¥{t.get('exit_price', 0):.2f}",
                f"{pnl:+.2f}%",
                t.get('shares', 0),
                t.get('hold_days', 0),
                str(t.get('reason_entry', ''))[:60],
                str(t.get('reason_exit', ''))[:60],
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row + i, column=col, value=v)
                _style_data_cell(cell, alt)
                if col == 6 and pnl > 0:
                    cell.font = Font(color="27AE60", bold=True)
                elif col == 6 and pnl < 0:
                    cell.font = Font(color="E74C3C", bold=True)
                if col in (1, 2):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        row += len(trades_list)
    else:
        row = _write_section_title(ws, row, " 交易明细（无交易）", max_col)
        row += 1

    # ---- 信号汇总 ----
    row += 1
    row = _write_section_title(ws, row, " 当前交易信号", max_col)
    row += 1

    if signal:
        current_signal_data = [
            ("信号动作", signal.action),
            ("建议仓位", f"{signal.position_size * 100:.0f}%"),
            ("入场价格", f"¥{signal.entry_price:.2f}"),
            ("止损价格", f"¥{signal.stop_loss:.2f}" if signal.stop_loss > 0 else "—"),
            ("目标价格", f"¥{signal.take_profit:.2f}" if signal.take_profit > 0 else "—"),
            ("信号理由", signal.reason),
        ]
        for i, (label, value) in enumerate(current_signal_data):
            alt = i % 2 == 1
            c1 = ws.cell(row=row + i, column=1, value=label)
            c1.font = LABEL_FONT
            _style_data_cell(c1, alt)
            c1.alignment = Alignment(horizontal="right", vertical="center")
            ws.merge_cells(start_row=row + i, start_column=2, end_row=row + i, end_column=max_col)
            c2 = ws.cell(row=row + i, column=2, value=value)
            _style_data_cell(c2, alt)
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        c = ws.cell(row=row, column=1, value="无交易信号")
        c.font = VALUE_FONT

    _auto_width(ws)


# ======================== Sheet 4: 关键结构 ========================

def _build_structure_sheet(ws, analyzer):
    """构建关键结构页面（中枢、笔、当前K线）"""
    max_col = 6
    current_price = analyzer.klines[-1].close if analyzer.klines else 0

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title = ws.cell(row=1, column=1, value="缠论关键结构数据")
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # ---- 中枢列表 ----
    zs_list = analyzer.zhongshus
    row = 3
    row = _write_section_title(ws, row, f" 中枢（共{len(zs_list)}个）", max_col)
    row += 1

    if zs_list:
        zs_headers = ["序号", "起始日", "结束日", "ZG(上沿)", "ZD(下沿)", "包含当前价"]
        for col, h in enumerate(zs_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_ACCENT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        row += 1

        for i, zs in enumerate(zs_list):
            alt = i % 2 == 1
            contains_price = "是" if zs.zd <= current_price <= zs.zg else \
                             ("上方" if current_price > zs.zg else "下方")
            vals = [i + 1, zs.start_date, zs.end_date,
                    f"¥{zs.zg:.2f}", f"¥{zs.zd:.2f}", contains_price]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row + i, column=col, value=v)
                _style_data_cell(cell, alt)
                if col in (2, 3):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        row += len(zs_list)
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        ws.cell(row=row, column=1, value="无中枢数据").font = VALUE_FONT
        row += 1

    # ---- 笔列表 ----
    row += 1
    bi_list = analyzer.bis
    row = _write_section_title(ws, row, f" 笔（共{len(bi_list)}条，显示最近30条）", max_col)
    row += 1

    if bi_list:
        bi_headers = ["序号", "方向", "起始日", "结束日", "起始价", "结束价"]
        for col, h in enumerate(bi_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        row += 1

        # 只显示最近30条
        display_bis = bi_list[-30:]
        for i, bi in enumerate(display_bis):
            alt = i % 2 == 1
            dir_cn = "↑ 上升" if bi.direction == "up" else "↓ 下降"
            dir_fill = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid") if bi.direction == "up" else PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
            vals = [i + 1, dir_cn, bi.start_date, bi.end_date,
                    f"¥{bi.start_price:.2f}", f"¥{bi.end_price:.2f}"]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row + i, column=col, value=v)
                _style_data_cell(cell, alt)
                if col == 2:
                    cell.fill = dir_fill
                if col in (3, 4):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    row += min(len(bi_list) if bi_list else 0, 30) + 1

    # ---- 最新K线 ----
    row += 1
    row = _write_section_title(ws, row, " 最新K线数据（最近10根）", max_col)
    row += 1

    klines = analyzer.klines[-10:] if analyzer.klines and len(analyzer.klines) > 10 else analyzer.klines
    if klines:
        k_headers = ["日期", "开盘", "收盘", "最高", "最低", "成交量"]
        for col, h in enumerate(k_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        row += 1

        for i, k in enumerate(klines):
            alt = i % 2 == 1
            vals = [k.date, f"¥{k.open:.2f}", f"¥{k.close:.2f}",
                    f"¥{k.high:.2f}", f"¥{k.low:.2f}", k.volume if hasattr(k, 'volume') and k.volume else ""]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row + i, column=col, value=v)
                _style_data_cell(cell, alt)
                if col == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    _auto_width(ws)
