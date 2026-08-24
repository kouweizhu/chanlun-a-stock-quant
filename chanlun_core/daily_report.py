#!/usr/bin/env python
"""
持仓日报生成器
读取 Excel 持仓表 → 运行缠论分析 → 生成 Markdown 日报
包含：现价/盈亏/止损告警/止盈提示/缠论信号/MACD状态

用法: python daily_report.py [--date YYYY-MM-DD]
"""
import sys, os, json, subprocess, argparse
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

# ============ 配置 ============
EXCEL_PATH = "D:/常用文件/持仓监控/持仓股票.xlsx"
REPORT_DIR = Path("D:/常用文件/持仓监控/收盘检测报告")
CHANLUN_DIR = Path(r"D:/常用文件/DeepSeek Harness项目/trading-skills/chanlun_core")

# 颜色/状态标记
CHECK = "✅"
WARN = "⚠️"
ALERT = "🔴"
OK = "🟢"

# ============ 数据读取 ============
def read_positions(excel_path):
    """读取 Excel 持仓表，返回 list[dict]"""
    wb = load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    positions = []
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # 跳过空行（以股票代码列为准）
        code_cell = row[0]
        if code_cell.value is None:
            continue
        # 确保代码是 6 位字符串
        code = str(code_cell.value).zfill(6)
        
        pos = {
            "code": code,
            "name": str(row[1].value) if row[1].value else "",
            "buy_date": row[2].value,  # datetime
            "buy_price": float(row[3].value) if row[3].value else 0,
            "reason": str(row[4].value) if row[4].value else "",
            "stop_loss": float(row[5].value) if row[5].value else None,
            "take_profit": float(row[6].value) if row[6].value else None,
        }
        positions.append(pos)
    
    return positions


def run_chanlun(code):
    """运行 quick_chanlun.py，返回解析后的 dict"""
    try:
        result = subprocess.run(
            ["python", str(CHANLUN_DIR / "quick_chanlun.py"), code],
            capture_output=True, text=True, timeout=120,
            cwd=str(CHANLUN_DIR)
        )
        text = result.stdout
        # 提取 JSON 部分
        start = text.find('{')
        end = text.rfind('}') + 1
        if start == -1 or end <= start:
            return None
        data = json.loads(text[start:end])
        return data.get("daily", {})
    except Exception as e:
        print(f"  [{code}] 缠论分析失败: {e}", file=sys.stderr)
        return None


def format_buy_date(d):
    """datetime → 'MM-DD' 字符串"""
    if d is None:
        return "?"
    if isinstance(d, datetime):
        return d.strftime("%m-%d")
    return str(d)[:10][5:]  # fallback


def pnl_pct(buy_price, current_price):
    """盈亏百分比"""
    if buy_price == 0:
        return 0
    return (current_price - buy_price) / buy_price * 100


def check_alerts(pos, current_price):
    """检查止损/止盈告警"""
    alerts = []
    status = OK
    
    # 止损检查
    if pos.get("stop_loss") and current_price <= pos["stop_loss"]:
        gap_pct = (current_price - pos["stop_loss"]) / pos["stop_loss"] * 100
        alerts.append(f"🔴 触发止损！现价{current_price:.2f} ≤ 止损{pos['stop_loss']:.2f}({gap_pct:+.1f}%)")
        status = ALERT
    elif pos.get("stop_loss") and current_price <= pos["stop_loss"] * 1.03:
        gap_pct = (current_price / pos["stop_loss"] - 1) * 100
        alerts.append(f"⚠️ 接近止损(距{gap_pct:.1f}%)：止损价{pos['stop_loss']:.2f}")
        if status != ALERT:
            status = WARN
    
    # 止盈检查
    if pos.get("take_profit") and current_price >= pos["take_profit"]:
        gap_pct = (current_price - pos["take_profit"]) / pos["take_profit"] * 100
        alerts.append(f"🎯 触发止盈！现价{current_price:.2f} ≥ 止盈{pos['take_profit']:.2f}(+{gap_pct:.1f}%)")
        if status != ALERT:
            status = OK  # 止盈是好事
    elif pos.get("take_profit") and current_price >= pos["take_profit"] * 0.95:
        gap_pct = (current_price / pos["take_profit"] - 1) * 100
        alerts.append(f"📈 接近止盈(距{abs(gap_pct):.1f}%)：止盈价{pos['take_profit']:.2f}")
    
    return status, alerts


def macd_summary(macd):
    """MACD 状态摘要"""
    if not macd:
        return "?"
    trend = macd.get("macd_trend", "?")
    cross = macd.get("dif_dea", "?")
    if cross == "golden_cross":
        cross_txt = "金叉"
    elif cross == "dead_cross":
        cross_txt = "死叉"
    else:
        cross_txt = cross
    trend_txt = "↑" if trend == "up" else "↓" if trend == "down" else trend
    return f"{cross_txt} {trend_txt}"


def recent_signals(buy_sell_points, days=60):
    """统计近期买卖点数量"""
    from datetime import timedelta
    cutoff = date.today() - timedelta(days=days)
    buys = 0
    sells = 0
    for bp in buy_sell_points:
        bp_date = datetime.strptime(bp["date"], "%Y-%m-%d").date()
        if bp_date >= cutoff:
            if bp["type"] == "buy":
                buys += 1
            else:
                sells += 1
    return buys, sells


# ============ 报告生成 ============
def generate_report(positions, chanlun_data, report_date):
    """生成 Markdown 报告"""
    lines = []
    lines.append(f"# 持仓日报 — {report_date}")
    lines.append("")
    lines.append(f"> 监控 {len(positions)} 只持仓股 | 自动生成 | 数据源: Baostock+缠论")
    lines.append("")
    
    # === 概览 ===
    lines.append("## 📊 概览")
    lines.append("")
    lines.append("| 状态 | 数量 |")
    lines.append("|------|:----:|")
    
    # 统计告警
    alerts_total = 0
    ok_list = []
    warn_list = []
    alert_list = []
    for p in positions:
        if p["code"] in chanlun_data:
            cl = chanlun_data[p["code"]]
            price = cl.get("current_price", 0)
            status, pos_alerts = check_alerts(p, price)
            if status == OK:
                ok_list.append(p)
            elif status == WARN:
                warn_list.append(p)
            else:
                alert_list.append(p)
            alerts_total += len(pos_alerts)
        else:
            ok_list.append(p)
    
    lines.append(f"| 🟢 正常 | {len(ok_list)} |")
    if warn_list:
        lines.append(f"| ⚠️ 关注 | {len(warn_list)} |")
    if alert_list:
        lines.append(f"| 🔴 告警 | {len(alert_list)} |")
    lines.append(f"| **总告警** | **{alerts_total}** |")
    lines.append("")
    
    # === 持仓明细 ===
    lines.append("## 💼 持仓明细")
    lines.append("")
    lines.append("| 代码 | 名称 | 买入价 | 现价 | 盈亏% | 止损 | 止盈 | 笔方向 | MACD | 买卖点(60d) | 状态 |")
    lines.append("|------|------|:-----:|:----:|:-----:|:----:|:----:|:------:|:----:|:-----------:|:----:|")
    
    for p in positions:
        code = p["code"]
        name = p["name"]
        buy_price = p["buy_price"]
        stop_loss = p.get("stop_loss")
        take_profit = p.get("take_profit")
        
        if code not in chanlun_data:
            lines.append(f"| {code} | {name} | {buy_price:.2f} | - | - | {stop_loss or '-'} | {take_profit or '-'} | ? | ? | ? | ⚫ 无数据 |")
            continue
        
        cl = chanlun_data[code]
        current_price = cl.get("current_price", 0)
        pnl = pnl_pct(buy_price, current_price)
        
        # 笔方向
        last_5 = cl.get("last_5_bis", [])
        bi_dir = last_5[-1]["direction"] if last_5 else "?"
        bi_arrow = "↑" if bi_dir == "up" else "↓" if bi_dir == "down" else "?"
        
        # MACD
        macd = cl.get("macd_status", {})
        macd_text = macd_summary(macd)
        
        # 买卖点
        bps = cl.get("buy_sell_points", [])
        buys, sells = recent_signals(bps)
        sig_text = f"买{buys}卖{sells}"
        
        # 状态
        status, alerts = check_alerts(p, current_price)
        
        sl_text = f"**{stop_loss:.2f}**" if stop_loss else "-"
        tp_text = f"**{take_profit:.2f}**" if take_profit else "-"
        
        lines.append(
            f"| {code} | {name} | {buy_price:.2f} | {current_price:.2f} | "
            f"{pnl:+.1f}% | {sl_text} | {tp_text} | {bi_arrow} | "
            f"{macd_text} | {sig_text} | {status} |"
        )
    
    lines.append("")
    
    # === 告警详情 ===
    all_alerts = []
    for p in positions:
        if p["code"] in chanlun_data:
            cl = chanlun_data[p["code"]]
            price = cl.get("current_price", 0)
            status, pos_alerts = check_alerts(p, price)
            for a in pos_alerts:
                all_alerts.append(f"- **{p['name']}({p['code']})**: {a}")
    
    if all_alerts:
        lines.append("## 🚨 告警详情")
        lines.append("")
        lines.extend(all_alerts)
        lines.append("")
    
    # === 各股信号 ===
    lines.append("## 📈 缠论信号")
    lines.append("")
    for p in positions:
        name = p["name"]
        code = p["code"]
        if code not in chanlun_data:
            lines.append(f"- **{name}({code})**: 数据获取失败")
            continue
        
        cl = chanlun_data[code]
        last_5 = cl.get("last_5_bis", [])
        bi_dir = last_5[-1]["direction"] if last_5 else "?"
        
        # 最新中枢
        zss = cl.get("zhongshus", [])
        latest_zs = zss[-1] if zss else None
        
        bps = cl.get("buy_sell_points", [])
        buys, sells = recent_signals(bps)
        current_price = cl.get("current_price", 0)
        
        lines.append(f"- **{name}({code})** @{current_price:.2f} | 笔方向:{bi_dir} | 近60天买{buys}卖{sells}")
        
        if latest_zs:
            zg, zd = latest_zs["zg"], latest_zs["zd"]
            in_zs = "中枢内" if zd <= current_price <= zg else \
                     "中枢上" if current_price > zg else "中枢下"
            lines.append(f"  中枢: [{zd:.2f}, {zg:.2f}] → {in_zs}")
    
    lines.append("")
    
    # === 止损止盈速查 ===
    lines.append("## 🎯 止损/止盈速查")
    lines.append("")
    lines.append("| 名称 | 止损价 | 现价 | 距止损 | 止盈价 | 距止盈 |")
    lines.append("|------|:-----:|:----:|:------:|:-----:|:------:|")
    
    for p in positions:
        code = p["code"]
        name = p["name"]
        sl = p.get("stop_loss")
        tp = p.get("take_profit")
        
        if code not in chanlun_data:
            continue
        
        cl = chanlun_data[code]
        current_price = cl.get("current_price", 0)
        
        sl_dist = f"{(current_price/sl-1)*100:+.1f}%" if sl else "-"
        tp_dist = f"{(current_price/tp-1)*100:+.1f}%" if tp else "-"
        sl_text = f"{sl:.2f}" if sl else "-"
        tp_text = f"{tp:.2f}" if tp else "-"
        
        lines.append(f"| {name} | {sl_text} | {current_price:.2f} | {sl_dist} | {tp_text} | {tp_dist} |")
    
    lines.append("")
    lines.append("---")
    lines.append(f"*报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
    lines.append(f"*下次运行: 下一个交易日 15:30*")
    
    return "\n".join(lines)


# ============ Main ============
def main():
    parser = argparse.ArgumentParser(description="生成持仓日报")
    parser.add_argument("--date", type=str, default=date.today().isoformat(),
                       help="报告日期 (YYYY-MM-DD)")
    args = parser.parse_args()
    report_date = args.date
    
    print(f"=== 持仓日报生成器 ===")
    print(f"日期: {report_date}")
    print()
    
    # 1. 读取持仓
    print("[1/3] 读取持仓表...")
    positions = read_positions(EXCEL_PATH)
    print(f"  共 {len(positions)} 只持仓: {', '.join(p['name'] for p in positions)}")
    
    # 2. 运行缠论分析
    print("\n[2/3] 运行缠论分析...")
    chanlun_data = {}
    for p in positions:
        code = p["code"]
        print(f"  {code} {p['name']}...", end=" ", flush=True)
        cl = run_chanlun(code)
        if cl:
            chanlun_data[code] = cl
            print(f"OK (现价={cl.get('current_price', '?')})")
        else:
            print("FAIL")
    
    # 3. 生成报告
    print("\n[3/3] 生成报告...")
    report = generate_report(positions, chanlun_data, report_date)
    
    # 保存
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = REPORT_DIR / f"{report_date}_持仓日报.md"
    report_path.write_text(report, encoding="utf-8")
    
    print(f"\n=== 报告已保存 ===")
    print(f"路径: {report_path}")
    print(f"大小: {len(report)} 字符")
    
    # 输出报告
    print("\n" + "="*60)
    print(report)


if __name__ == "__main__":
    main()
