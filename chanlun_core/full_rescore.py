#!/usr/bin/env python
"""full_rescore.py — 全部 120 只候选股 Tavily 重评 + 生成全新报告"""

import sys, os, json, time, shutil
from date_utils import date_to_str, parse_date_to_datetime
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from file_utils import safe_read_json
from config_loader import A500_COMPOSITE_THRESHOLD
from pool_screener import scan_news, W_TECH, W_FUND, W_NEWS, NEWS_TOP_N
from composite_scorer import compute_3d_score, position_reason
from quick_fundamental import classify_by_industry

CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".scanner_cache.json")
OUTPUT_BASE = "D:/常用文件/股票池推荐股"
TOP_N = 10   # 报告中展示前N只

os.makedirs(OUTPUT_BASE, exist_ok=True)

# ============================================================
# 1. 加载候选股 + 真实评分
# ============================================================
cache = safe_read_json(CACHE)
candidates = cache['candidates']
print(f"加载 {len(candidates)} 只候选股")

# 尝试加载 pool_screener Phase 2 的真实技术/基本面评分
PHASE2_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".phase2_results.json")
real_scores = {}
if os.path.exists(PHASE2_JSON):
    phase2_data = safe_read_json(PHASE2_JSON)
    for s in phase2_data:
        real_scores[s['code']] = {
            'tech_score': s.get('tech_score', 0),
            'fund_score': s.get('fund_score', 0),
            'roe': s.get('roe', 0),
            'np_margin': s.get('np_margin', 0),
            'pe': s.get('pe', 0),
            'pb': s.get('pb', 0),
            'industry': s.get('industry', ''),
        }
    print(f"加载 {len(real_scores)} 只真实技术/基本面评分\n")
else:
    print("⚠ 未找到 .phase2_results.json，将使用估分（请先运行 pool_screener.py）\n")

# ============================================================
# 2. 先计算技术+基本面预评分（消息面=50中性），再对 Top N 做消息面扫描
# ============================================================
scored = []
for c in candidates:
    rs = real_scores.get(c['code'], {})
    tech_score = rs.get('tech_score')
    fund_score = rs.get('fund_score')

    # 无真实分时回退到估分
    if tech_score is None:
        tech_score = {5: 86, 3: 65}.get(int(c['score']), 50)
    if fund_score is None:
        fund_score = 70

    # 预评分（消息面=50）
    pre_3d = compute_3d_score(
        tech_score=tech_score, fund_score=fund_score, news_score=50,
        w_tech=W_TECH, w_fund=W_FUND, w_news=W_NEWS, resonance_penalty=True,
    )
    scored.append({
        'code': c['code'], 'name': c['name'],
        'price': c.get('price', 0), 'score': c['score'],
        'pattern': c.get('pattern', ''), 'buy_type': c.get('buy_type', ''),
        'buy_date': c.get('buy_date', ''), 'buy_price': c.get('buy_price', 0),
        'zg': c.get('zg', 0), 'zd': c.get('zd', 0),
        'total_bis': c.get('total_bis', 0), 'total_zs': c.get('total_zs', 0),
        'buy_count': c.get('buy_count', 0), 'sell_count': c.get('sell_count', 0),
        'tech_score': tech_score, 'fund_score': fund_score,
        'news_score': 50.0, 'news_detail': '跳过(非Top30)',
        'composite': pre_3d.composite, 'grade': pre_3d.grade,
        'position': pre_3d.position, 'position_pct': f"{pre_3d.position*100:.0f}%",
        'can_buy': pre_3d.can_buy, 'reason': position_reason(pre_3d),
        'resonance': pre_3d.components.get('resonance_penalty_applied', False),
        'roe': rs.get('roe', 0), 'np_margin': rs.get('np_margin', 0),
        'pe': rs.get('pe', 0), 'pb': rs.get('pb', 0),
        'industry': rs.get('industry', ''),
    })

scored.sort(key=lambda s: -s['composite'])
print(f"技术+基本面预评分完成: {len(scored)} 只\n")

# ── 消息面扫描：仅 Top N ──
top_n = min(NEWS_TOP_N, len(scored))
print(f"消息面扫描 Top {top_n}...")
news_updated = 0
for i in range(top_n):
    s = scored[i]
    print(f"  [{i+1}/{top_n}] {s['code']} {s['name']}...", end=" ", flush=True)
    try:
        news_score, news_detail = scan_news(s['code'], s['name'])
        s['news_score'] = news_score
        s['news_detail'] = news_detail
        result_3d = compute_3d_score(
            tech_score=s['tech_score'], fund_score=s['fund_score'],
            news_score=news_score,
            w_tech=W_TECH, w_fund=W_FUND, w_news=W_NEWS,
            resonance_penalty=True,
        )
        s['composite'] = result_3d.composite
        s['grade'] = result_3d.grade
        s['position'] = result_3d.position
        s['position_pct'] = f"{result_3d.position*100:.0f}%"
        s['can_buy'] = result_3d.can_buy
        s['reason'] = position_reason(result_3d)
        s['resonance'] = result_3d.components.get('resonance_penalty_applied', False)
        news_updated += 1
        print(f"news={news_score:.0f} → composite={result_3d.composite:.0f}({result_3d.grade})")
    except Exception as e:
        print(f"失败: {e}")
    if i < top_n - 1:
        time.sleep(3.0)  # Tavily rate limit

print(f"  消息面扫描完成: {news_updated}/{top_n}\n")

# 排序
scored.sort(key=lambda s: -s['composite'])

# ============================================================
# 3. 保存全量结果
# ============================================================
date_str = datetime.now().strftime('%Y-%m-%d')
result_path = os.path.join(OUTPUT_BASE, f"全量评分_{date_str}.json")
with open(result_path, 'w', encoding='utf-8') as f:
    json.dump(scored, f, ensure_ascii=False, indent=2)
print(f"\n全量结果: {result_path}")

# ============================================================
# 4. 打印 Top N（过滤综合分 < COMPOSITE_THRESHOLD 的） 
# 4. 打印 Top N（过滤综合分 < A500_COMPOSITE_THRESHOLD 的） 
COMPOSITE_THRESHOLD = A500_COMPOSITE_THRESHOLD  # 从 config 读取
top_raw = scored[:TOP_N]
# 过滤掉低于阈值的（保持三维系统一致性）
top = [s for s in top_raw if s['composite'] >= COMPOSITE_THRESHOLD]
skipped_low = len(top_raw) - len(top)
if skipped_low > 0:
    print(f"\n⚠ {skipped_low} 只综合分 < {COMPOSITE_THRESHOLD}，已从 Top 10 剔除")
    # 从后面补充
    for s in scored[TOP_N:]:
        if s['composite'] >= COMPOSITE_THRESHOLD and len(top) < TOP_N:
            top.append(s)
        if len(top) >= TOP_N:
            break
print(f"\n{'='*80}")
print(f"🏆 Top {len(top)} (Tavily 消息面, 近一周)")
print(f"{'='*80}")
print(f"{'#':>3} {'代码':<8} {'名称':<8} {'综合':>5} {'等级':>4} {'技术':>5} {'基本面':>5} {'消息':>5} {'仓位':>5} {'模式':<30}")
for i, s in enumerate(top, 1):
    icon = "🟢" if s['grade'] == 'A' else ("🟡" if s['grade'] == 'B' else "🟠")
    print(f"{i:>3} {s['code']:<8} {s['name']:<8} {s['composite']:>5.1f} {icon}{s['grade']:>3} "
          f"{s['tech_score']:>5.1f} {s['fund_score']:>5.1f} {s['news_score']:>5.1f} {s['position_pct']:>5} "
          f"{s['pattern'][:28]}")

# ============================================================
# 5. 生成 MD 总表
# ============================================================
md_lines = [
    f"# A500 股票池智能筛选汇总",
    f"**扫描日期**: {date_str} | **消息面**: Tavily Search API (近一周)",
    f"**候选股**: {len(scored)} 只 | **筛选**: 30天买点=5分/潜在结构=3分 | >30天剔除",
    "",
    "## 🏆 推荐列表（按综合分降序）",
    "",
    "| # | 代码 | 名称 | 综合 | 等级 | 技术 | 基本面 | 消息 | 仓位 | 买点模式 |",
    "|---|------|------|:----:|:----:|:----:|:------:|:----:|:----:|------|",
]
for i, s in enumerate(top, 1):
    icon = "🟢" if s['grade'] == 'A' else ("🟡" if s['grade'] == 'B' else "🟠")
    md_lines.append(
        f"| {i} | {s['code']} | {s['name']} | **{s['composite']}** | "
        f"{icon}{s['grade']} | {s['tech_score']} | {s['fund_score']} | "
        f"{s['news_score']} | {s['position_pct']} | {s['pattern'][:25]} |"
    )

md_lines += [
    "",
    "## 📰 消息面详情",
    "",
    "| # | 名称 | 消息分 | 关键词统计 | 判断 |",
    "|---|------|:------:|------|------|",
]
for i, s in enumerate(top, 1):
    judgment = "利好" if s['news_score'] > 55 else ("利空" if s['news_score'] < 45 else "中性")
    md_lines.append(f"| {i} | {s['name']} | {s['news_score']} | {s['news_detail']} | {judgment} |")

md_lines += [
    "",
    "---",
    "*报告由 Hermes 三维分析系统自动生成*",
    "*技术面: 缠论日线分析 | 基本面: Baostock财务数据 | 消息面: Tavily Search API*",
]

md_path = os.path.join(OUTPUT_BASE, f"扫描汇总_{date_str}.md")
with open(md_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(md_lines))
print(f"MD 总表: {md_path}")

# ============================================================
# 6. 生成 Excel 总表
# ============================================================
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    xlsx_path = os.path.join(OUTPUT_BASE, f"扫描汇总_{date_str}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top10"

    headers = ['排名', '代码', '名称', '综合分', '等级', '技术', '基本面', '消息',
               '仓位', '现价', '模式', '买点类型', '买点日期', '买点价', 'ZG', 'ZD',
               '消息详情', '可建仓']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    grade_fills = {
        'A': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'B': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'C': PatternFill(start_color='F4B4C2', end_color='F4B4C2', fill_type='solid'),
    }

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

    for row_idx, s in enumerate(scored[:10], 2):
        vals = [row_idx-1, s['code'], s['name'], s['composite'], s['grade'],
                s['tech_score'], s['fund_score'], s['news_score'],
                s['position_pct'], s['price'], s['pattern'][:30],
                s['buy_type'], s['buy_date'], s['buy_price'],
                s['zg'], s['zd'], s['news_detail'], '是' if s['can_buy'] else '否']
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center')
        gf = grade_fills.get(s['grade'])
        if gf:
            ws.cell(row=row_idx, column=2).fill = gf
            ws.cell(row=row_idx, column=5).fill = gf

    widths = [5,8,10,8,5,7,8,7,6,9,28,8,14,9,9,9,22,6]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # Sheet 2: 全部候选
    ws2 = wb.create_sheet("全部候选")
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border
    for row_idx, s in enumerate(scored, 2):
        vals = [row_idx-1, s['code'], s['name'], s['composite'], s['grade'],
                s['tech_score'], s['fund_score'], s['news_score'],
                s['position_pct'], s['price'], s['pattern'][:30],
                s['buy_type'], s['buy_date'], s['buy_price'],
                s['zg'], s['zd'], s['news_detail'], '是' if s['can_buy'] else '否']
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row_idx, column=col, value=val)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center')
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    wb.save(xlsx_path)
    print(f"Excel 总表: {xlsx_path}")
except ImportError:
    print("Excel 跳过 (openpyxl 未安装)")

# ============================================================
# 7. 生成个股 MD 评分报告（Top 40）
# ============================================================
print(f"\n生成个股报告...")
for i, s in enumerate(top):
    stock_dir = os.path.join(OUTPUT_BASE, f"{s['name']}_{s['code']}")
    os.makedirs(stock_dir, exist_ok=True)
    md_file = os.path.join(stock_dir, f"{s['code']}_score_report.md")

    content = f"""# {s['name']}({s['code']}) 三维系统评分报告

**生成日期**: {datetime.now().strftime('%Y-%m-%d %H:%M')}
**现价**: ¥{s['price']}

---

## 📊 综合评分: {s['composite']} ({s['grade']}级)

| 维度 | 得分 | 权重 | 加权 |
|------|:----:|:----:|:----:|
| 技术面 | {s['tech_score']} | 40% | {s['tech_score']*0.4:.1f} |
| 基本面 | {s['fund_score']} | 30% | {s['fund_score']*0.3:.1f} |
| 消息面 | {s['news_score']} | 30% | {s['news_score']*0.3:.1f} |
| **综合** | **{s['composite']}** | 100% | — |

> 等级: A≥70 B≥60 C≥50 | {'共振惩罚已应用' if s.get('resonance') else '无共振惩罚'}

---

## 💰 仓位建议: {s['position_pct']}

{s['reason']}

---

## 🔍 技术面详情

| 项目 | 内容 |
|------|------|
| 扫描模式 | {s['pattern']} |
| 买点类型 | {s['buy_type']} |
| 买点日期 | {s['buy_date']} |
| 买点价格 | ¥{s['buy_price']} |
| 最近中枢 | ZG=¥{s['zg']}, ZD=¥{s['zd']} |
| 笔数/中枢数 | {s['total_bis']}/{s['total_zs']} |
| 技术得分 | {s['tech_score']} |

---

## 📰 消息面 (Tavily 近一周)

- 消息得分: {s['news_score']}/100
- 扫描结果: {s['news_detail']}

---

## ⚠️ 风险提示

- 本报告由三维分析系统自动生成，仅供参考，不构成投资建议
- 技术面基于日线缠论分析，未做30分钟多级别确认
- 消息面为自动化搜索摘要，可能遗漏重要信息
- 投资决策请结合个人风险承受能力和市场整体环境
"""
    with open(md_file, 'w', encoding='utf-8') as f:
        f.write(content)

    if (i+1) % 10 == 0:
        print(f"  MD {i+1}/{len(top)} 完成")

# ============================================================
# 8. 生成 Top 10 HTML 技术分析报告
# ============================================================
top10 = scored[:10]
print(f"\n生成 Top 10 HTML 缠论分析图...")
html_ok = 0
for i, s in enumerate(top10):
    stock_dir = os.path.join(OUTPUT_BASE, f"{s['name']}_{s['code']}")
    os.makedirs(stock_dir, exist_ok=True)
    html_path = os.path.join(stock_dir, f"{s['code']}_chanlun_analysis.html")

    # 如果 HTML 已有且更新（今天生成的），跳过
    if os.path.exists(html_path):
        mtime = os.path.getmtime(html_path)
        if (time.time() - mtime) < 86400:  # 24小时内
            print(f"  [{i+1}/10] {s['name']}({s['code']}) HTML 已有(今日) → 跳过")
            html_ok += 1
            continue

    print(f"  [{i+1}/10] {s['name']}({s['code']}) 生成HTML...", end=" ", flush=True)
    try:
        from data_manager import DataManager
        from generate_analysis import RecursiveTimingSystem, HTMLVisualizer
        dm = DataManager()
        rec = RecursiveTimingSystem(dm)
        daily = rec.run_full_analysis(s['code'])
        if daily and daily.klines:
            viz = HTMLVisualizer(s['code'], s['name'], daily)
            viz.generate_html(html_path)
            print("✓")
            html_ok += 1
        else:
            print("✗ 分析失败")
    except Exception as e:
        print(f"✗ {str(e)[:40]}")

print(f"\n{'='*60}")
print(f"✅ 全部完成!")
print(f"  Top 10 HTML: {html_ok}/10 份")
print(f"  MD 个股报告: {len(top)} 份")
print(f"  MD 总表: {md_path}")
print(f"  Excel 总表: {xlsx_path}")
print(f"  全量数据: {result_path}")
